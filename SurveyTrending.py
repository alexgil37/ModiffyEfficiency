import openpyxl
import re
import os
import sys
import xlsxwriter
import statistics
import sympy


def main(path, savePath):
    grossTotalCounts = list()
    backgroundCounts = list()
    removableCounts = list()
    invalidSheets = list()
    netCPMRem = list()
    netActRem = list()
    netActTotal = list()
    netCPMTotal = list()
    invalidFiles = list()
    allNetAct = list()
    allRemAct = list()
    descriptionList = list()

    # Create the output folder
    if not os.path.isdir(savePath):
        os.mkdir(savePath)

    # create excel QC file
    QCworkbook = xlsxwriter.Workbook(savePath + '\\' + 'BetaGammaTrending.xlsx')
    QCworksheet = QCworkbook.add_worksheet("All Values")
    StatisticSheet = QCworkbook.add_worksheet("Stats per Sheet")

    # create columns with headers
    QCworksheet.write(0, 0, 'File Name')
    QCworksheet.write(0, 1, 'Survey Number')
    QCworksheet.write(0, 2, 'Date')
    QCworksheet.write(0, 3, 'Survey Tech')
    QCworksheet.write(0, 4, 'Count room Tech')
    QCworksheet.write(0, 5, 'Date of Count Room')
    QCworksheet.write(0, 6, 'Survey Type')
    QCworksheet.write(0, 7, 'Level of Posting')
    QCworksheet.write(0, 8, 'Item Surveyed')
    QCworksheet.write(0, 9, 'No.')
    QCworksheet.write(0, 10, 'Description/Location')
    QCworksheet.write(0, 11, 'Measurement Description')
    QCworksheet.write(0, 12, 'Total Activity Instrument Efficiency')
    QCworksheet.write(0, 13, 'Gross counts of Total Activity')
    QCworksheet.write(0, 14, 'Background counts of Total activity')
    QCworksheet.write(0, 15, 'Net Activity of Total Activity')
    QCworksheet.write(0, 16, 'Removable Instrument Efficiency')
    QCworksheet.write(0, 17, 'Gross Counts of Removable')
    QCworksheet.write(0, 18, 'Net Activity of Removable')

    # Create statistics sheet Headers
    StatisticSheet.write(0, 0, 'File Name')
    StatisticSheet.write(0, 1, 'Survey Number')
    StatisticSheet.write(0, 2, 'Sheet Name')
    StatisticSheet.write(0, 3, 'Total Activity Min')
    StatisticSheet.write(0, 4, 'Total Activity Max')
    StatisticSheet.write(0, 5, 'Total Activity Average')
    StatisticSheet.write(0, 6, 'Total Activity Standard Deviation')
    StatisticSheet.write(0, 7, 'Removable Min')
    StatisticSheet.write(0, 8, 'Removable Max')
    StatisticSheet.write(0, 9, 'Removable Average')
    StatisticSheet.write(0, 10, 'Removable Standard Deviation')

    StatisticSheet.write(0, 13, 'Overall Total Activity  Minimum')
    StatisticSheet.write(0, 14, 'Overall Total Activity Maximum')
    StatisticSheet.write(0, 15, 'Overall Total Activity Average')
    StatisticSheet.write(0, 16, 'Overall Total Activity Standard Deviation')
    StatisticSheet.write(0, 17, 'Overall Removable Minimum')
    StatisticSheet.write(0, 18, 'Overall Removable Maximum')
    StatisticSheet.write(0, 20, 'Overall Removable Average')
    StatisticSheet.write(0, 21, 'Overall Removable Standard Deviation')

    def resource_path(relative_path):
        base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.dirname(__file__)))
        return os.path.join(base_path, relative_path)

    def getListOfFiles(dirName):
        listOfFile = os.listdir(dirName)
        allFiles = list()

        for file in listOfFile:
            fullPath = os.path.join(dirName, file)
            if os.path.isdir(fullPath):
                allFiles = allFiles + getListOfFiles(fullPath)
            else:
                allFiles.append(fullPath)

        return allFiles

    def find_cell(currentSheet, parameterToFind):
        for row in range(1, 40):
            for column in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":  # Here you can add or reduce the columns

                cell = "{}{}".format(column, row)

                if currentSheet[cell].value == parameterToFind:
                    return [row, column, currentSheet[cell]]

        return [0, 0, None]

    def find_cell_with_keyword(currentSheet, keyword):
        for row in range(5, 40):
            for column in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":  # Here you can add or reduce the columns

                cell = "{}{}".format(column, row)

                if keyword in str(currentSheet[cell].value):
                    return [row, column, currentSheet[cell]]

        return [0, 0, None]

    def find_title_vals(currentSheet):
        # find Survey number
        sampleIdTitleCell = find_cell(currentSheet, "Survey No")
        if sampleIdTitleCell[2] is None:
            sampleIdTitleCell = find_cell(currentSheet, "Survey Number")
        sampleIdCell = find_title_data(currentSheet, sampleIdTitleCell)

        # find survey techs
        surveyTechTitleCell = find_cell(currentSheet, "Survey Tech")
        surveyTechCell = find_title_data(currentSheet, surveyTechTitleCell)

        # find Count room tech
        countRoomTechTitleCell = find_cell(currentSheet, "Count Room Tech")
        countRoomTechCell = find_title_data(currentSheet, countRoomTechTitleCell)

        # find type
        typeTitleCell = find_cell(currentSheet, "Survey Type")
        if typeTitleCell[2] is None:
            typeCell = None
        else:
            typeCell = find_title_data(currentSheet, typeTitleCell)

        # find Level of Posting
        postTitleCell = find_cell(currentSheet, "Level Of Posting")
        if postTitleCell[2] is None:
            postTitleCell = find_cell(currentSheet, "Level of Posting")
        postingCell = find_title_data(currentSheet, postTitleCell)

        # find Item Surveyed
        locationTitleCell = find_cell(currentSheet, "Item Surveyed")
        if locationTitleCell[2] is None:
            locationTitleCell = find_cell(currentSheet, "Survey Number")
        locationCell = find_title_data(currentSheet, locationTitleCell)

        return [sampleIdCell, surveyTechCell, countRoomTechCell, typeCell, postingCell, locationCell]

    def find_title_data(currentSheet, titleCell):
        row = int(titleCell[0])
        col = chr(ord(titleCell[1]) + 1)

        while type(currentSheet[col + str(row)]).__name__ == 'MergedCell':
            col = chr(ord(col) + 1)

        valueCell = currentSheet[col + str(row)]

        return valueCell

    def check_for_BettaGamma(num):
        found = 0
        for row in range(1, 30):
            for column in "GHIJKLMNOPQRSTUVWXYZ":

                modelVal = currentSheet[column + str(row)].value
                if modelVal != "Beta-Gamma":
                    continue

                found += 1
                if found == num:
                    return [row, column]

        return [None, None]

    def checkForMap():
        for row in range(1, 10):
            for column in "ABCDEFGHIJ":
                modelCell = "{}{}".format(column, row)
                if currentSheet[modelCell].value is not None:
                    return False

        return True

    files = getListOfFiles(path)

    percentFormat = QCworkbook.add_format({'num_format': '0.00%'})

    QCfileRow = 1
    SecondSheetRow = 1
    # Create Date format to write to xlsx files
    dateFormat = QCworkbook.add_format({'num_format': 'mm/dd/yyyy'})

    for file in files:
        try:
            # For Openpyxl
            theFile = openpyxl.load_workbook(file)
        except:
            continue
        allSheetNames = theFile.sheetnames
        print(file)
        print("All sheet names {} ".format(theFile.sheetnames))

        for x in allSheetNames:
            badFile = False
            print("Current sheet name is {}".format(x))
            currentSheet = theFile[x]

            # If it is a map sheet skip
            currentSheetString = str(currentSheet)
            currentSheetString = currentSheetString[12:]
            checkBlankString = currentSheetString[0:5]
            currentSheetString = currentSheetString[0:3]
            if currentSheetString == "Map" or checkBlankString == "Blank":
                continue

            # Check which format it is
            betaRow, betaCol = check_for_BettaGamma(3)
            index = 0
            needToContinue = False

            # Prevent repeats be added to invalid Sheets list
            if betaRow is None or betaCol is None:
                needToContinue = True
                if len(invalidSheets) > 0:
                    if invalidSheets.count(file) == 0:
                        invalidSheets.append(file)

                else:
                    invalidSheets.append(file)

            # Find MDC Value and DPM for total activity as well as Removable DPM
            else:
                backgroundCounts.clear()
                grossTotalCounts.clear()
                removableCounts.clear()

                removableBetaRow, removableBetaCol = check_for_BettaGamma(4)
                countsCol = betaCol
                backgroundCol = chr(ord(betaCol) + 1)
                removableCountsCol = removableBetaCol

                n = 1
                # Go until it is not None
                while currentSheet[betaCol + str(betaRow + n)].value is None:
                    n += 1

                descriptionCell = find_cell_with_keyword(currentSheet, "Description")

                descriptionColumn = descriptionCell[1]
                descriptionRow = descriptionCell[0]
                descriptionOffset = 0

                # There will always be at most 20 counts per survey
                for cell in range(n + 1, n + 21):
                    descriptionOffset += 1
                    descriptionValue = currentSheet[descriptionColumn + str(descriptionOffset + descriptionRow)].value
                    cellValue = currentSheet[countsCol + str(betaRow + cell)].value
                    backgroundValue = currentSheet[backgroundCol + str(betaRow + cell)].value
                    if backgroundValue is None:
                        backgroundValue = 0
                    removableValue = currentSheet[removableCountsCol + str(removableBetaRow + cell)].value

                    # If there is nothing in both removable and total activity
                    if cellValue is None and removableValue is None:
                        continue

                    # If there is removable but no total activity
                    elif cellValue is None and removableValue is not None:
                        backgroundCounts.append(None)
                        grossTotalCounts.append(None)
                        removableCounts.append(removableValue)
                        descriptionList.append(descriptionValue)

                    # If there is total activity but no removable
                    elif cellValue is not None and removableValue is None:
                        backgroundCounts.append(backgroundValue)
                        grossTotalCounts.append(cellValue)
                        removableCounts.append(None)
                        descriptionList.append(descriptionValue)

                    # If there is both removable and total activity
                    else:
                        backgroundCounts.append(backgroundValue)
                        grossTotalCounts.append(cellValue)
                        removableCounts.append(removableValue)
                        descriptionList.append(descriptionValue)

                    index += 1

            if needToContinue is True:
                continue

            # Find totalEfficiency
            efficiencyRow, efficiencyCol = check_for_BettaGamma(1)
            efficiencyRow += 4
            totalEfficiency = currentSheet[efficiencyCol + str(efficiencyRow)].value

            ttemp, efficiencyCol = check_for_BettaGamma(2)
            remEfficiency = currentSheet[efficiencyCol + str(efficiencyRow)].value

            # Find Removable Background
            efficiencyRow, efficiencyCol = check_for_BettaGamma(2)
            efficiencyRow += 8
            bkgRem = currentSheet[efficiencyCol + str(efficiencyRow)].value
            if bkgRem is None:
                bkgRem = 0

            # Find Correction Factor
            efficiencyRow, efficiencyCol = check_for_BettaGamma(1)
            efficiencyRow += 5
            correctionFactor = currentSheet[efficiencyCol + str(efficiencyRow)].value
            correctionFactor = correctionFactor

            # ***********Find Title Data**********
            titleVals = find_title_vals(currentSheet)

            # find date
            dateTitleCell = find_cell(currentSheet, "Date")
            if dateTitleCell[2] is None or dateTitleCell[1] == 0:
                dateTitleCell = find_cell(currentSheet, "Date Counted")
            dateCell = find_title_data(currentSheet, dateTitleCell)

            # Find the Name of the worksheet
            currentSheetString = str(currentSheet)
            currentSheetString = currentSheetString[12:]
            currentSheetString = currentSheetString[:-2]

            netCPMRem.clear()
            netActRem.clear()
            netActTotal.clear()
            netCPMTotal.clear()

            # Find DPMs
            badFile = False
            for i in range(0, len(removableCounts)):

                if removableCounts[i] is None:
                    netCPMRem.append(None)
                    netActRem.append(None)

                elif type(bkgRem) != str:
                    if remEfficiency is None:
                        netCPMRem.append(None)
                        netActRem.append(None)
                    else:
                        netCPMRem.append(removableCounts[i] - (bkgRem / 60))
                        netActRem.append(netCPMRem[i] / remEfficiency)

                else:
                    try:
                        bkgRem = bkgRem[1:]
                        bkgRem = eval(bkgRem)

                        # Better security but we need to test much more
                        # bkgRem = int(sympy.sympify(bkgRem))

                        if correctionFactor is None:
                            netCPMRem.append(None)
                            netActRem.append(None)
                        else:
                            netCPMRem.append((removableCounts[i] - (bkgRem / 60)) * correctionFactor)
                            netActRem.append(netCPMRem[i] / remEfficiency)

                    except:
                        if invalidFiles.count(file) == 0:
                            invalidFiles.append(file)
                            badFile = True

            # total activity calculations
            for i in range(0, len(grossTotalCounts)):
                if grossTotalCounts[i] is None:
                    netCPMTotal.append(None)
                    netActTotal.append(None)

                elif type(bkgRem) != str:
                    try:
                        netCPMTotal.append(grossTotalCounts[i] - (backgroundCounts[i]))
                        netActTotal.append(netCPMTotal[i] / totalEfficiency * correctionFactor)
                    except:
                        if invalidFiles.count(file) == 0:
                            invalidFiles.append(file)
                            badFile = True
                        else:
                            invalidFiles.append(file)
                            badFile = True

                else:
                    try:
                        backgroundCounts[i] = backgroundCounts[i][1:]
                        backgroundCounts[i] = eval(backgroundCounts[i])

                        # Better security but we need to test much more
                        # backgroundCounts[i] = int(sympy.sympify(backgroundCounts[i]))

                        netCPMTotal.append(grossTotalCounts[i] - (backgroundCounts[i]) * correctionFactor)
                        netActTotal.append(netCPMTotal[i] / totalEfficiency)

                    except:
                        if invalidFiles.count(file) == 0:
                            invalidFiles.append(file)
                            badFile = True

            # Write the results to the QC file
            # Write the current Worksheet
            head, tail = os.path.split(file)
            if badFile is True:
                continue

            for i in range(0, len(removableCounts)):
                QCworksheet.write(QCfileRow, 0, tail)  # File Name
                QCworksheet.write(QCfileRow, 1, titleVals[0].value)  # Survey Number
                QCworksheet.write(QCfileRow, 2, dateCell.value, dateFormat)  # Date
                QCworksheet.write(QCfileRow, 3, titleVals[1].value)  # Survey Tech
                QCworksheet.write(QCfileRow, 4, titleVals[2].value)  # Count room Tech
                if titleVals[3] is not None:
                    QCworksheet.write(QCfileRow, 6, titleVals[3].value)  # Survey Type
                try:
                    QCworksheet.write(QCfileRow, 7, titleVals[4].value)  # Level of Posting
                    QCworksheet.write(QCfileRow, 8, titleVals[5].value)  # Item Surveyed
                    QCworksheet.write(QCfileRow, 9, descriptionList[i])
                    QCworksheet.write(QCfileRow, 10, totalEfficiency, percentFormat)  # totalEfficiency
                    QCworksheet.write(QCfileRow, 11, round(grossTotalCounts[i]))  # Gross Counts Total
                    QCworksheet.write(QCfileRow, 12, round(backgroundCounts[i]))  # Background total activity
                    QCworksheet.write(QCfileRow, 13, round(netActTotal[i]))  # DPM total activity
                    QCworksheet.write(QCfileRow, 14, remEfficiency, percentFormat)  # Removable instrument Efficeincy
                    QCworksheet.write(QCfileRow, 15, round(removableCounts[i]))  # Gross removable Counts
                    QCworksheet.write(QCfileRow, 16, round(netActRem[i]))  # Removable DPM
                except:
                    QCworksheet.write(QCfileRow, 7, titleVals[4].value)  # Level of Posting
                    QCworksheet.write(QCfileRow, 8, titleVals[5].value)  # Item Surveyed
                    QCworksheet.write(QCfileRow, 9, descriptionList[i])  # Item Surveyed
                    QCworksheet.write(QCfileRow, 10, totalEfficiency, percentFormat)  # totalEfficiency
                    QCworksheet.write(QCfileRow, 11, grossTotalCounts[i])  # Gross Counts Total
                    QCworksheet.write(QCfileRow, 12, backgroundCounts[i])  # Background total activity
                    QCworksheet.write(QCfileRow, 13, netActTotal[i])  # DPM total activity
                    QCworksheet.write(QCfileRow, 14, remEfficiency, percentFormat)  # Removable instrument Efficeincy
                    QCworksheet.write(QCfileRow, 15, removableCounts[i])  # Gross removable Counts
                    QCworksheet.write(QCfileRow, 16, netActRem[i])  # Removable DPM

                QCfileRow += 1

            allNetAct.extend(netActTotal)
            allRemAct.extend(netActRem)

            # Find the statistics
            netActTotal = list(filter(None, netActTotal))
            netActRem = list(filter(None, netActRem))

            if len(netActTotal) != 0:
                totalAverage = sum(netActTotal) / len(netActTotal)
                totalMax = max(netActTotal)
                totalMin = min(netActTotal)
                totalStdDev = statistics.pstdev(netActTotal)
            else:
                totalAverage = None
                totalMax = None
                totalMin = None
                totalStdDev = None

            if len(netActRem) != 0:
                removableAvg = sum(netActRem) / len(netActRem)
                removableMax = max(netActRem)
                removableMin = min(netActRem)
                removableStdDev = statistics.pstdev(netActRem)
            else:
                removableAvg = None
                removableMax = None
                removableMin = None
                removableStdDev = None

            try:
                StatisticSheet.write(SecondSheetRow, 0, tail)  # File Name
                StatisticSheet.write(SecondSheetRow, 1, titleVals[0].value)  # Survey Number
                StatisticSheet.write(SecondSheetRow, 2, currentSheetString)  # Current Sheet
                StatisticSheet.write(SecondSheetRow, 3, round(totalMin))
                StatisticSheet.write(SecondSheetRow, 4, round(totalMax))
                StatisticSheet.write(SecondSheetRow, 5, round(totalAverage))
                StatisticSheet.write(SecondSheetRow, 6, totalStdDev)
                StatisticSheet.write(SecondSheetRow, 7, round(removableMin))
                StatisticSheet.write(SecondSheetRow, 8, round(removableMax))
                StatisticSheet.write(SecondSheetRow, 9, round(removableAvg))
                StatisticSheet.write(SecondSheetRow, 10, removableStdDev)
            except:
                StatisticSheet.write(SecondSheetRow, 0, tail)  # File Name
                StatisticSheet.write(SecondSheetRow, 1, titleVals[0].value)  # Survey Number
                StatisticSheet.write(SecondSheetRow, 2, currentSheetString)  # Current Sheet
                StatisticSheet.write(SecondSheetRow, 3, totalMin)
                StatisticSheet.write(SecondSheetRow, 4, totalMax)
                StatisticSheet.write(SecondSheetRow, 5, totalAverage)
                StatisticSheet.write(SecondSheetRow, 6, totalStdDev)
                StatisticSheet.write(SecondSheetRow, 7, removableMin)
                StatisticSheet.write(SecondSheetRow, 8, removableMax)
                StatisticSheet.write(SecondSheetRow, 9, removableAvg)
                StatisticSheet.write(SecondSheetRow, 10, removableStdDev)

            SecondSheetRow += 1

        theFile.close()

    """redo but for the other list"""
    print("IN THE INVALID BETA ")
    del grossTotalCounts
    del backgroundCounts
    del removableCounts
    grossTotalCounts = list()
    removableCounts = list()

    for file in invalidSheets:
        theFile = openpyxl.load_workbook(file)
        allSheetNames = theFile.sheetnames
        print(file)
        print("All sheet names {} ".format(theFile.sheetnames))

        for x in allSheetNames:
            print("Current sheet name is {}".format(x))
            currentSheet = theFile[x]
            badFile = False

            # If it is a map sheet skip
            currentSheetString = str(currentSheet)
            currentSheetString = currentSheetString[12:]
            checkBlankString = currentSheetString[0:5]
            currentSheetString = currentSheetString[0:3]
            if currentSheetString == "Map" or checkBlankString == "Blank":
                continue

            betaRow, betaCol = check_for_BettaGamma(3)
            if betaRow is not None or betaCol is not None:
                continue

            if checkForMap():
                continue

            betaRow, betaCol = check_for_BettaGamma(1)
            if betaRow is None or betaCol is None:
                continue

            else:
                grossTotalCounts.clear()
                removableCounts.clear()

                betaRow, betaCol = check_for_BettaGamma(1)
                removableBetaRow, removableBetaCol = check_for_BettaGamma(2)

                n = 1
                # Go until it is not None
                while currentSheet[betaCol + str(betaRow + n)].value != "gross counts":
                    n += 1

                descriptionCell = find_cell_with_keyword(currentSheet, "Description")

                descriptionColumn = descriptionCell[1]
                descriptionRow = descriptionCell[0]
                descriptionOffset = 0

                # There will always be at most 20 counts per survey
                for cell in range(n + 1, n + 21):
                    descriptionOffset += 1
                    descriptionValue = currentSheet[descriptionColumn + str(descriptionOffset + descriptionRow)].value
                    cellValue = currentSheet[betaCol + str(betaRow + cell)].value
                    removableValue = currentSheet[removableBetaCol + str(removableBetaRow + cell)].value

                    # If there is nothing in both removable and total activity
                    if cellValue is None and removableValue is None:
                        continue

                    # If there is removable but no total activity
                    elif cellValue is None and removableValue is not None:
                        grossTotalCounts.append(None)
                        removableCounts.append(removableValue)
                        descriptionList.append(descriptionValue)

                    # If there is total activity but no removable
                    elif cellValue is not None and removableValue is None:
                        grossTotalCounts.append(cellValue)
                        removableCounts.append(None)
                        descriptionList.append(descriptionValue)

                    # If there is both removable and total activity
                    else:
                        grossTotalCounts.append(cellValue)
                        removableCounts.append(removableValue)
                        descriptionList.append(descriptionValue)

                    index += 1

            # Find efficiency
            efficiencyRow, efficiencyCol = check_for_BettaGamma(1)
            efficiencyRow += 4
            totalEfficiency = currentSheet[efficiencyCol + str(efficiencyRow)].value

            ttemp, efficiencyCol = check_for_BettaGamma(2)
            remEfficiency = currentSheet[efficiencyCol + str(efficiencyRow)].value

            # Find Background
            efficiencyRow, efficiencyCol = check_for_BettaGamma(2)
            efficiencyRow += 5
            bkgRem = currentSheet[efficiencyCol + str(efficiencyRow)].value

            efficiencyRow, efficiencyCol = check_for_BettaGamma(1)
            efficiencyRow += 5
            bkgTotal = currentSheet[efficiencyCol + str(efficiencyRow)].value

            # ***********Find Title Data**********
            titleVals = find_title_vals(currentSheet)

            # find date
            dateTitleCell = find_cell(currentSheet, "Date")
            if dateTitleCell[2] is None or dateTitleCell[1] == 0:
                dateTitleCell = find_cell(currentSheet, "Date Counted")
            dateCell = find_title_data(currentSheet, dateTitleCell)

            currentSheetString = str(currentSheet)
            currentSheetString = currentSheetString[12:]
            currentSheetString = currentSheetString[:-2]

            netCPMRem.clear()
            netActRem.clear()
            netActTotal.clear()
            netCPMTotal.clear()

            # Find DPMs
            for i in range(0, len(removableCounts)):
                try:
                    if removableCounts[i] is None:
                        netCPMRem.append(None)
                        netActRem.append(None)
                    else:
                        netCPMRem.append(removableCounts[i] - (bkgRem / 60))
                        netActRem.append(netCPMRem[i] / remEfficiency)

                except:
                    try:
                        if removableCounts[i] is None:
                            netCPMRem.append(None)
                            netActRem.append(None)
                        else:
                            netCPMRem.append(removableCounts[i])
                            netActRem.append(netCPMRem[i] / remEfficiency)

                    except:
                        if invalidFiles.count(file) == 0:
                            invalidFiles.append(file)
                            badFile = True

            # total activity calculations

            for i in range(0, len(grossTotalCounts)):
                try:
                    if grossTotalCounts[i] is None:
                        netCPMTotal.append(None)
                        netActTotal.append(None)
                    else:
                        netCPMTotal.append(grossTotalCounts[i] - (bkgTotal))
                        netActTotal.append(netCPMTotal[i] / totalEfficiency)
                except:
                    try:
                        if grossTotalCounts[i] is None:
                            netCPMTotal.append(None)
                            netActTotal.append(None)
                        else:
                            netCPMTotal.append(grossTotalCounts[i])
                            netActTotal.append(netCPMTotal[i] / totalEfficiency)
                    except:
                        if invalidFiles.count(file) == 0:
                            invalidFiles.append(file)
                            badFile = True

            if badFile is True:
                continue


            head, tail = os.path.split(file)
            for i in range(0, len(removableCounts)):
                QCworksheet.write(QCfileRow, 0, tail)  # File Name
                QCworksheet.write(QCfileRow, 1, titleVals[0].value)  # Survey Number
                QCworksheet.write(QCfileRow, 2, dateCell.value, dateFormat)  # Date
                QCworksheet.write(QCfileRow, 3, titleVals[1].value)  # Survey Tech
                QCworksheet.write(QCfileRow, 4, titleVals[2].value)  # Count room Tech
                # QCworksheet.write(QCfileRow, 5, secondDateCell.value, dateFormat)  # Date of Count Room
                try:
                    QCworksheet.write(QCfileRow, 6, titleVals[3].value)  # Survey Type
                    QCworksheet.write(QCfileRow, 7, titleVals[4].value)  # Level of Posting
                    QCworksheet.write(QCfileRow, 8, titleVals[5].value)  # Item Surveyed
                    QCworksheet.write(QCfileRow, 9, descriptionList[i])  # Item Surveyed
                    QCworksheet.write(QCfileRow, 10, totalEfficiency, percentFormat)  # totalEfficiency
                    QCworksheet.write(QCfileRow, 11, round(grossTotalCounts[i]))  # Gross Counts Total
                    QCworksheet.write(QCfileRow, 12, round(bkgTotal))  # Background total activity
                    QCworksheet.write(QCfileRow, 13, round(netActTotal[i])) # DPM total activity
                    QCworksheet.write(QCfileRow, 14, remEfficiency, percentFormat)  # Removable instrument Efficeincy
                    QCworksheet.write(QCfileRow, 15, round(removableCounts[i]))  # Gross removable Counts
                    QCworksheet.write(QCfileRow, 16, round(netActRem[i]))  # Removable DPM
                except:
                    QCworksheet.write(QCfileRow, 6, titleVals[3].value)  # Survey Type
                    QCworksheet.write(QCfileRow, 7, titleVals[4].value)  # Level of Posting
                    QCworksheet.write(QCfileRow, 8, titleVals[5].value)  # Item Surveyed
                    QCworksheet.write(QCfileRow, 9, descriptionList[i])  # Item Surveyed
                    QCworksheet.write(QCfileRow, 10, totalEfficiency, percentFormat)  # totalEfficiency
                    QCworksheet.write(QCfileRow, 11, grossTotalCounts[i])  # Gross Counts Total
                    QCworksheet.write(QCfileRow, 12, bkgTotal)  # Background total activity
                    QCworksheet.write(QCfileRow, 13, netActTotal[i])  # DPM total activity
                    QCworksheet.write(QCfileRow, 14, remEfficiency, percentFormat)  # Removable instrument Efficeincy
                    QCworksheet.write(QCfileRow, 15, removableCounts[i])  # Gross removable Counts
                    QCworksheet.write(QCfileRow, 16, netActRem[i])  # Removable DPM

                QCfileRow += 1

            # Find the statistics
            allNetAct.extend(netActTotal)
            allRemAct.extend(netActRem)

            # Find the statistics
            netActTotal = list(filter(None, netActTotal))
            netActRem = list(filter(None, netActRem))

            if len(netActTotal) != 0:
                totalAverage = sum(netActTotal) / len(netActTotal)
                totalMax = max(netActTotal)
                totalMin = min(netActTotal)
                totalStdDev = statistics.pstdev(netActTotal)
            else:
                totalAverage = None
                totalMax = None
                totalMin = None
                totalStdDev = None

            if len(netActRem) != 0:
                removableAvg = sum(netActRem) / len(netActRem)
                removableMax = max(netActRem)
                removableMin = min(netActRem)
                removableStdDev = statistics.pstdev(netActRem)
            else:
                removableAvg = None
                removableMax = None
                removableMin = None
                removableStdDev = None

            try:
                StatisticSheet.write(SecondSheetRow, 0, tail)  # File Name
                StatisticSheet.write(SecondSheetRow, 1, titleVals[0].value)  # Survey Number
                StatisticSheet.write(SecondSheetRow, 2, currentSheetString)  # Current Sheet
                StatisticSheet.write(SecondSheetRow, 3, round(totalMin))
                StatisticSheet.write(SecondSheetRow, 4, round(totalMax))
                StatisticSheet.write(SecondSheetRow, 5, round(totalAverage))
                StatisticSheet.write(SecondSheetRow, 6, totalStdDev)
                StatisticSheet.write(SecondSheetRow, 7, round(removableMin))
                StatisticSheet.write(SecondSheetRow, 8, round(removableMax))
                StatisticSheet.write(SecondSheetRow, 9, round(removableAvg))
                StatisticSheet.write(SecondSheetRow, 10, removableStdDev)
            except:
                StatisticSheet.write(SecondSheetRow, 0, tail)  # File Name
                StatisticSheet.write(SecondSheetRow, 1, titleVals[0].value)  # Survey Number
                StatisticSheet.write(SecondSheetRow, 2, currentSheetString)  # Current Sheet
                StatisticSheet.write(SecondSheetRow, 3, totalMin)
                StatisticSheet.write(SecondSheetRow, 4, totalMax)
                StatisticSheet.write(SecondSheetRow, 5, totalAverage)
                StatisticSheet.write(SecondSheetRow, 6, totalStdDev)
                StatisticSheet.write(SecondSheetRow, 7, removableMin)
                StatisticSheet.write(SecondSheetRow, 8, removableMax)
                StatisticSheet.write(SecondSheetRow, 9, removableAvg)
                StatisticSheet.write(SecondSheetRow, 10, removableStdDev)

            SecondSheetRow += 1

            grossTotalCounts.clear()
            removableCounts.clear()

        theFile.close()
        theFile.save(file)
        grossTotalCounts.clear()
        removableCounts.clear()

    # Find overall stats
    allNetAct = list(filter(None, allNetAct))
    allRemAct = list(filter(None, allRemAct))

    try:
        allTotalAverage = sum(allNetAct) / len(allNetAct)
        allTotalMax = max(allNetAct)
        allTotalMin = min(allNetAct)
        allTotalStdDev = statistics.pstdev(allNetAct)
    except:
        allTotalAverage = None
        allTotalMax = None
        allTotalMin = None
        allTotalStdDev = None

    try:
        allRemovableAvg = sum(allRemAct) / len(allRemAct)
        allRemovableMax = max(allRemAct)
        allRemovableMin = min(allRemAct)
        allRemovableStdDev = statistics.pstdev(allRemAct)
    except:
        allRemovableAvg = None
        allRemovableMax = None
        allRemovableMin = None
        allRemovableStdDev = None

    try:
        StatisticSheet.write(1, 13, round(allTotalMin))
        StatisticSheet.write(1, 14, round(allTotalMax))
        StatisticSheet.write(1, 15, round(allTotalAverage))
        StatisticSheet.write(1, 16, allTotalStdDev)
        StatisticSheet.write(1, 17, round(allRemovableMin))
        StatisticSheet.write(1, 18, round(allRemovableMax))
        StatisticSheet.write(1, 19, round(allRemovableAvg))
        StatisticSheet.write(1, 20, allRemovableStdDev)
    except:
        StatisticSheet.write(1, 13, allTotalMin)
        StatisticSheet.write(1, 14, allTotalMax)
        StatisticSheet.write(1, 15, allTotalAverage)
        StatisticSheet.write(1, 16, allTotalStdDev)
        StatisticSheet.write(1, 17, allRemovableMin)
        StatisticSheet.write(1, 18, allRemovableMax)
        StatisticSheet.write(1, 19, allRemovableAvg)
        StatisticSheet.write(1, 20, allRemovableStdDev)

    if len(invalidFiles) > 0:
        FailedSheet = QCworkbook.add_worksheet("Statistics")
        FailedSheet.write(0, 0, 'Invalid Files')
        x = 1
        for file in invalidFiles:
            FailedSheet.write(x, 0, file)
            x += 1

    QCworkbook.close()
    os.startfile(savePath + '\\' + 'BetaGammaTrending.xlsx')
