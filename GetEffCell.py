import openpyxl
import json
import os
import xlsxwriter


def main(path, savePath):
    if os.path.isfile(savePath):
        os.mkdir(savePath)

    # create excel QC file
    QCworkbook = xlsxwriter.Workbook(savePath + '\\' + 'QC.xlsx')
    QCworksheet = QCworkbook.add_worksheet()

    # create columns with headers
    QCworksheet.write(0, 0, 'File Name')
    QCworksheet.write(0, 1, 'Sheet Name')
    QCworksheet.write(0, 2, 'Survey Number')
    QCworksheet.write(0, 3, 'Instrument S/N')
    QCworksheet.write(0, 4, 'Cal DueDate')
    QCworksheet.write(0, 5, 'Old Efficiency')
    QCworksheet.write(0, 6, 'New Efficiency')



    def getListOfFiles(dirName):
        listOfFile = os.listdir(dirName)
        allFiles = list()

        for file in listOfFile:
            fullPath = os.path.join(dirName, file)
            if os.path.isdir(fullPath):
                allFiles = allFiles + getListOfFiles(fullPath)
            else:
                allFiles.append(fullPath)

        print(allFiles)

        return allFiles


    def find_instrument_model_cell(currentSheet):
        for row in range(1, 50):
            for column in "ABCDEFGHIJKLMNOPQRSTUV":  # Here you can add or reduce the columns
                modelCell = "{}{}".format(column, row)
                if currentSheet[modelCell].value == instrumentModel:
                    print("the row is {0} and the column {1}" .format(row, column))
                    print(currentSheet[modelCell].value)
                    print(modelCell)

                    return [row, column, modelCell]

        return [0, 0, None]


    def find_instrument_sn_cell(instModelRow, instModelColumn):
        snRow = str(int(instModelRow) + 1)
        snCol = instModelColumn
        snCell = currentSheet[snCol + snRow]

        print('xxxxxxxxxxx')
        print(snRow)
        print(snCol)
        print(snCell.value)

        return snCell


    def find_cal_due_date(instModelRow, instModelColumn):
        calRow = str(int(instModelRow) + 2)
        calCol = chr(ord(instModelColumn))
        calCell = currentSheet[calCol + calRow]

        print(calCell.value)
        print(calRow)
        print(calCol)

        return calCell


    def find_instrument_efficiency(instModelRow, instModelColumn):
        effRow = str(int(instModelRow) + 3)
        effCol = chr(ord(instModelColumn) + 2)
        effCell = currentSheet[effCol + effRow]

        if type(effCell).__name__ == 'MergedCell':
            effCol = chr(ord(instModelColumn) + 3)
            effCell = currentSheet[effCol + effRow]

        print(effCell.value)
        print(effRow)
        print(effCol)

        return effCell

    """
    Take the current sheet and returns the value of the cell to the right of the cell containing Survey No.
    """
    def find_survey_number(currentSheet):
        for column in "ABCDEFGHIJKLMNOPQRSTUV":
            for row in range(1, 50):
                modelCell = "{}{}".format(column, row)
                newCol = column
                value = currentSheet[modelCell].value

                while (currentSheet[modelCell].value == "Survey No") or (currentSheet[modelCell].value == "Survey Number"):
                    if newCol >= "V":
                        continue

                    newCol = chr(ord(newCol) + 1)
                    cell = currentSheet[newCol + str(row)]

                    if type(cell).__name__ != 'MergedCell':
                        newVal = currentSheet[newCol + str(row)].value
                        test = 0
                        return newVal


    def find_efficiency(instSNcell, instEfficiencyCell):
        for inst in instrumentsData:
            if inst['sn'] == instSNcell.value:
                return [inst['sn'], inst['betaEfficiency']]

        return [None, None]


    instrumentModel = '2360/43-93'
    #instrumentId = '227413/PR295918'

    filesWithNoMatchingSN = list()
    sheetsOfFilesWithNoMatchingSN = list()

    files = getListOfFiles(path)

    """This is used for the exe"""
    # with open(file=resource_path("package.json")) as instruments_file:
    #     instrumentsData = json.load(instruments_file)

    """This is used when running the program"""
    with open('package.json') as instruments_file:
        instrumentsData = json.load(instruments_file)

    QCfileRow = 1
    dateFormat = QCworkbook.add_format({'num_format': 'mm/dd/yyyy'})

    for file in files:
        theFile = openpyxl.load_workbook(file)
        allSheetNames = theFile.sheetnames

        print("All sheet names {} ".format(theFile.sheetnames))

        for x in allSheetNames:
            print("Current sheet name is {}" .format(x))
            currentSheet = theFile[x]
            instModelRow = find_instrument_model_cell(currentSheet)[0]
            instModelColumn = find_instrument_model_cell(currentSheet)[1]
            instModelCell = find_instrument_model_cell(currentSheet)[2]
            surveyNumber = find_survey_number(currentSheet)

            print("The cell is {}, the row is {} and the column is {} ".format(instModelCell, instModelRow, instModelColumn))

            if instModelCell is None:
                continue
            instSNcell = find_instrument_sn_cell(instModelRow, instModelColumn)
            instEfficiencyCell = find_instrument_efficiency(instModelRow, instModelColumn)

            oldinstEfficiencyCell = instEfficiencyCell.value
            instCalDueDate = find_cal_due_date(instModelRow, instModelColumn)
            serialNumber = find_efficiency(instSNcell, instEfficiencyCell)

            if serialNumber[0] is None:
                filesWithNoMatchingSN.append(file)
                sheetsOfFilesWithNoMatchingSN.append(currentSheet)

            # Find the file name
            head, tail = os.path.split(file)

            # Find the Name of the worksheet
            currentSheetString = str(currentSheet)
            currentSheetString = currentSheetString[12:]
            currentSheetString = currentSheetString[:-2]

            # Write the current Worksheet
            QCworksheet.write(QCfileRow, 0, tail)
            QCworksheet.write(QCfileRow, 1, currentSheetString)
            QCworksheet.write(QCfileRow, 2, surveyNumber)
            QCworksheet.write(QCfileRow, 3, instSNcell.value)
            QCworksheet.write(QCfileRow, 4, instCalDueDate.value, dateFormat)
            QCworksheet.write(QCfileRow, 5, oldinstEfficiencyCell)
            QCworksheet.write(QCfileRow, 6, serialNumber[1])

            QCfileRow += 1

        theFile.close()
        theFile.save(file)

    QCworkbook.close()

    print("The files with no s/n are {}, the sheet is {}".format(filesWithNoMatchingSN, sheetsOfFilesWithNoMatchingSN))
    os.startfile(savePath + '\\' + 'QC.xlsx')

    return
