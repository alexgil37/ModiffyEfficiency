import openpyxl
import json
import os
import sys
import xlsxwriter
import totalActivityWeird

instrumentModel = '2360/43-93'

filesWithNoMatchingSN = list()
sheetsOfFilesWithNoMatchingSN = list()
counts = list()
backgroundCounts = list()
invalidSheets = list()
badfile = list()


def main(path, savePath):
    if os.path.isfile(savePath):
        os.mkdir(savePath)

    # create excel QC file
    QCworkbook = xlsxwriter.Workbook(savePath + '\\' + 'Total_Activity.xlsx')
    QCworksheet = QCworkbook.add_worksheet()

    # create columns with headers
    QCworksheet.write(0, 0, 'File Name')
    QCworksheet.write(0, 1, 'Sheet Name')
    QCworksheet.write(0, 2, 'Survey Number')
    QCworksheet.write(0, 3, 'Gross Counts')
    QCworksheet.write(0, 4, 'Background Counts')
    QCworksheet.write(0, 5, 'Efficiency')


    def resource_path(relative_path):
        base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.dirname(__file__)))
        return os.path.join(base_path, relative_path)


    def getListOfFiles(dirName):
        listOfFile = os.listdir(dirName)
        allFiles = list()

        for file in listOfFile:
            fullPath = os.path.join(dirName, file)
            if os.path.isdir(fullPath):
                allFiles = allFiles + getListOfFiles(fullPath)
            else:
                allFiles.append(fullPath)

        return allFiles


    def find_instrument_model_cell(currentSheet):
        for row in range(1, 30):
            for column in "GHIJKLMNOPQRSTUVWXYZ":  # Here you can add or reduce the columns
                modelCell = "{}{}".format(column, row)
                print(file)
                if currentSheet[modelCell].value == instrumentModel:
                    print("the row is {0} and the column {1}" .format(row, column))
                    print(currentSheet[modelCell].value)
                    print(modelCell)

                    return [row, column, modelCell]

        return [0, 0, None]


    def check_for_BettaGamma(num):
        found = 0
        for row in range(1, 30):
            for column in "GHIJKLMNOPQRSTUVWXYZ":

                modelVal = currentSheet[column +str(row)].value
                if modelVal != "Beta-Gamma":
                    continue

                found += 1
                if found == num:
                    return [row, column]

        return [None, None]


    def find_instrument_sn_cell(instModelRow, instModelColumn):
        snRow = str(int(instModelRow) + 1)
        snCol = instModelColumn
        snCell = currentSheet[snCol + snRow]

        return snCell


    def find_instrument_efficiency(instModelRow, instModelColumn):
        effRow = str(int(instModelRow) + 3)
        effCol = chr(ord(instModelColumn) + 2)
        effCell = currentSheet[effCol + effRow]

        if type(effCell).__name__ == 'MergedCell':
            effCol = chr(ord(instModelColumn) + 3)
            effCell = currentSheet[effCol + effRow]

        return effCell

    """
    Take the current sheet and returns the value of the cell to the right of the cell containing Survey No.
    """
    def find_survey_number(currentSheet):
        for column in "ABCDEFGHI":
            for row in range(1, 20):
                modelCell = "{}{}".format(column, row)
                newCol = column

                while (currentSheet[modelCell].value == "Survey No") or (currentSheet[modelCell].value == "Survey Number"):
                    if newCol >= "V":
                        continue

                    newCol = chr(ord(newCol) + 1)
                    cell = currentSheet[newCol + str(row)]

                    if type(cell).__name__ != 'MergedCell':
                        newVal = currentSheet[newCol + str(row)].value
                        return newVal

        return "None"


    def find_efficiency(instSNcell):
        for inst in instrumentsData:
            if inst['sn'] == instSNcell.value:
                return [inst['sn'], inst['betaEfficiency']]

        return [None, None]


    def find_backgroud():
        for row in range(1, 30):
            for column in "GHIJKLMNOPQRSTUVWXYZ":

                modelVal = currentSheet[column + str(row)].value
                if modelVal == "Background Counts":
                    Col = chr(ord(column) + 5)
                    cell = currentSheet[Col + str(row)]

                    # If merged cell go over 1
                    if type(cell).__name__ != 'MergedCell':
                        Col = chr(ord(column) + 1)
                        modelVal = currentSheet[Col + str(row)].value
                        if modelVal is not None:
                            return [0, 0, modelVal]

                    else:
                        modelVal = currentSheet[Col + str(row)].value
                        if modelVal is not None:
                            return [0, 0, modelVal]

                if modelVal == "Bldg Material Bkg":
                    return [row, column, None]

        return [None, None, None]


    files = getListOfFiles(path)

    """This is used for the exe"""
    # with open(file=resource_path("package.json")) as instruments_file:
    #     instrumentsData = json.load(instruments_file)

    """This is used when running the program"""
    with open('package.json') as instruments_file:
        instrumentsData = json.load(instruments_file)

    QCfileRow = 1

    for file in files:
        theFile = openpyxl.load_workbook(file)
        allSheetNames = theFile.sheetnames

        for x in allSheetNames:
            print("Current sheet name is {}" .format(x))
            currentSheet = theFile[x]
            instModelRow, instModelColumn, instModelCell = find_instrument_model_cell(currentSheet)
            surveyNumber = find_survey_number(currentSheet)

            if instModelCell is None:
                continue
            instSNcell = find_instrument_sn_cell(instModelRow, instModelColumn)
            instEfficiencyCell = find_instrument_efficiency(instModelRow, instModelColumn)

            oldinstEfficiencyCell = instEfficiencyCell.value
            serialNumber = find_efficiency(instSNcell)

            if serialNumber[0] is None:
                filesWithNoMatchingSN.append(file)
                sheetsOfFilesWithNoMatchingSN.append(currentSheet)

            # Find the file name
            head, tail = os.path.split(file)

            # Find the Name of the worksheet
            currentSheetString = str(currentSheet)
            currentSheetString = currentSheetString[12:]
            currentSheetString = currentSheetString[:-2]

            betaRow, betaCol = check_for_BettaGamma(3)
            backgroundCol = chr(ord(betaCol) + 1)
            index = 0


            if betaRow is None or betaCol is None:
                invalidSheets.append(file)
                continue

            else:
                # There will always be at most 20 counts per survey
                n = 1
                # Go until you get to Gross Counts
                while currentSheet[betaCol + str(betaRow+n)].value is None:
                    n += 1

                for cell in range(n+1, n+21):
                    cellValue = currentSheet[betaCol + str(betaRow+cell)].value
                    if cellValue is None:
                        continue
                    else:

                        counts.append(cellValue)
                        backgroundValue = currentSheet[backgroundCol + str(betaRow+cell)].value
                        backgroundCounts.append(backgroundValue)
                        print(file)
                        print("cell Value: " + str(cellValue))
                        print("betaCol: " + betaCol)
                        print("backgroundCol: " + backgroundCol)
                        print("backgroundValue: " + str(backgroundValue))

                    index += 1

            for x in range(0, len(counts)):

                # Write the current Worksheet
                QCworksheet.write(QCfileRow, 0, tail)
                QCworksheet.write(QCfileRow, 1, currentSheetString)
                QCworksheet.write(QCfileRow, 2, surveyNumber)
                QCworksheet.write(QCfileRow, 3, counts[x])
                QCworksheet.write(QCfileRow, 4, backgroundCounts[x])
                QCworksheet.write(QCfileRow, 5, oldinstEfficiencyCell)

                QCfileRow += 1

            counts.clear()
            backgroundCounts.clear()

        theFile.close()
        theFile.save(file)

    for file in invalidSheets:
        theFile = openpyxl.load_workbook(file)
        allSheetNames = theFile.sheetnames

        for x in allSheetNames:
            print("Current sheet name is {}" .format(x))
            currentSheet = theFile[x]
            instModelRow, instModelColumn, instModelCell = find_instrument_model_cell(currentSheet)
            surveyNumber = find_instrument_model_cell(currentSheet)

            if instModelCell is None:
                continue
            instSNcell = find_instrument_model_cell(instModelRow)
            instEfficiencyCell = find_instrument_model_cell(instModelRow)

            oldinstEfficiencyCell = instEfficiencyCell.value
            serialNumber = find_instrument_model_cell(instSNcell)

            if serialNumber[0] is None:
                filesWithNoMatchingSN.append(file)
                sheetsOfFilesWithNoMatchingSN.append(currentSheet)

            # Find the file name
            head, tail = os.path.split(file)

            # Find the Name of the worksheet
            currentSheetString = str(currentSheet)
            currentSheetString = currentSheetString[12:]
            currentSheetString = currentSheetString[:-2]

            backgroundRow, backgroundColumn, backgroundVal = find_backgroud()
            betaRow, betaCol = check_for_BettaGamma(3)
            backgroundCol = chr(ord(instModelColumn) + 1)
            index = 0

            if backgroundRow is None and backgroundColumn is None and backgroundVal is None:
                badfile.append(file)
                continue

            #  If we can use the column for the background
            if backgroundVal is None:
                n = 1

                # Go until you get to Gross Counts
                while currentSheet[betaCol + str(betaRow + n)].value is None:
                    n += 1

                for cell in range(n + 1, n + 21):
                    cellValue = currentSheet[betaCol + str(betaRow + cell)].value
                    if cellValue is None:
                        continue
                    else:
                        counts.append(cellValue)

            # If we are given a single value for the background
            elif backgroundRow is None and backgroundColumn is None:
                n = 0

                # Go until you get to Gross Counts
                while currentSheet[betaCol + str(betaRow+n)].value is None:
                    n += 1

                for cell in range(n+1, n+21):
                    cellValue = currentSheet[betaCol + str(betaRow+cell)].value
                    if cellValue is None:
                        continue
                    else:
                        counts.append(cellValue)
                        backgroundCounts.append(backgroundVal)

                    index += 1

            for x in range(0, len(counts)):

                # Write the current Worksheet
                QCworksheet.write(QCfileRow, 0, tail)
                QCworksheet.write(QCfileRow, 1, currentSheetString)
                QCworksheet.write(QCfileRow, 2, surveyNumber)
                QCworksheet.write(QCfileRow, 3, counts[x])
                QCworksheet.write(QCfileRow, 4, backgroundCounts[x])
                QCworksheet.write(QCfileRow, 5, oldinstEfficiencyCell)

                QCfileRow += 1

            counts.clear()
            backgroundCounts.clear()

        theFile.close()
        theFile.save(file)

    QCworkbook.close()

    print(badfile)

    print("The files with no s/n are {}, the sheet is {}".format(filesWithNoMatchingSN, sheetsOfFilesWithNoMatchingSN))
    os.startfile(savePath + '\\' + 'Total_Activity.xlsx')

    return
