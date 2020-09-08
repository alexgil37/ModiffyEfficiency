import shutil
import shutil
import openpyxl
import json
import os
import sys
import xlsxwriter
from pycel import ExcelCompiler

def main(path, savePath):

    def getFile(dirName):
        listOfFile = os.listdir(dirName)
        allFiles = list()

        for file in listOfFile:
            filePath = os.path.join(dirName, file)
            theFile = openpyxl.load_workbook(filePath)

            break

        return theFile

    def write_to_file(currentLocationSheet, currentLocRow, row):

        currentLocationSheet.cell(currentLocRow, 1).value = dataFileWS[("{}{}".format("B", row))].value
        currentLocationSheet.cell(currentLocRow, 2).value = dataFileWS[("{}{}".format("C", row))].value
        currentLocationSheet.cell(currentLocRow, 3).value = dataFileWS[("{}{}".format("F", row))].value
        currentLocationSheet.cell(currentLocRow, 4).value = dataFileWS[("{}{}".format("G", row))].value
        currentLocationSheet.cell(currentLocRow, 5).value = dataFileWS[("{}{}".format("H", row))].value
        currentLocationSheet.cell(currentLocRow, 6).value = dataFileWS[("{}{}".format("I", row))].value

    def find_location(dataFileWS):

        # sheet current row
        sbCurrentRow = 3
        cwCurrentRow = 3
        bCurrentRow = 3
        firstSeCurrentRow = 3
        secondSeCurrentRow = 3
        secondHwCurrentRow = 3
        secondOutCurrentRow = 3
        thirdSeCurrentRow = 3
        fourthSeCurrentRow = 3
        fifthSeCurrentRow = 3
        sixthSeCurrentRow = 3
        seventhSeCurrentRow = 3
        seventh726CurrentRow = 3
        eigthSeCurrentRow = 3
        roofEastCurrentRow = 3
        roofWestCurrentRow = 3
        restCurrentRow = 1

        for column in "D":
            for row in range(1, 4000):

                cell1 = "{}{}".format(column, row)

                if dataFileWS[cell1].value is None:
                    break

                if "Sub-Basement" in dataFileWS[cell1].value:
                    sbCurrentRow = sbCurrentRow + 1
                    write_to_file(sbSheet, sbCurrentRow, row)

                elif "Basement" in dataFileWS[cell1].value:
                    bCurrentRow = bCurrentRow + 1
                    write_to_file(bSheet, bCurrentRow, row)

                elif "Catwalk" in dataFileWS[cell1].value:
                    cwCurrentRow = cwCurrentRow + 1
                    write_to_file(cwSheet, cwCurrentRow, row)

                elif "1st" in dataFileWS[cell1].value and "Service Elevator" in dataFileWS[cell1].value:
                    firstSeCurrentRow = firstSeCurrentRow + 1
                    write_to_file(firstSeSheet, firstSeCurrentRow, row)

                elif "2nd" in dataFileWS[cell1].value and "Service Elevator" in dataFileWS[cell1].value:
                    secondSeCurrentRow = secondSeCurrentRow + 1
                    write_to_file(secondSeSheet, secondSeCurrentRow, row)

                elif "2nd" in dataFileWS[cell1].value and "Hallway" in dataFileWS[cell1].value:
                    secondHwCurrentRow = secondHwCurrentRow + 1
                    write_to_file(secondHwSheet, secondHwCurrentRow, row)

                elif "2nd" in dataFileWS[cell1].value and "Out" in dataFileWS[cell1].value:
                    secondOutCurrentRow = secondOutCurrentRow + 1
                    write_to_file(secondOutSheet, secondOutCurrentRow, row)

                elif "3rd" in dataFileWS[cell1].value and "Service Elevator" in dataFileWS[cell1].value:
                    thirdSeCurrentRow = thirdSeCurrentRow + 1
                    write_to_file(thirdSeSheet, thirdSeCurrentRow, row)

                elif "4th" in dataFileWS[cell1].value and "Service Elevator" in dataFileWS[cell1].value:
                    fourthSeCurrentRow = fourthSeCurrentRow + 1
                    write_to_file(fourthSeSheet, fourthSeCurrentRow, row)

                elif "5th" in dataFileWS[cell1].value and "Service Elevator" in dataFileWS[cell1].value:
                    fifthSeCurrentRow = fifthSeCurrentRow + 1
                    write_to_file(fifthSeSheet, fifthSeCurrentRow, row)

                elif "6th" in dataFileWS[cell1].value and "Service Elevator" in dataFileWS[cell1].value:
                    sixthSeCurrentRow = sixthSeCurrentRow + 1
                    write_to_file(sixthSeSheet, sixthSeCurrentRow, row)

                elif "7th" in dataFileWS[cell1].value and "Service Elevator" in dataFileWS[cell1].value:
                    seventhSeCurrentRow = seventhSeCurrentRow + 1
                    write_to_file(seventhSeSheet, seventhSeCurrentRow, row)

                elif "7th" in dataFileWS[cell1].value and "726" in dataFileWS[cell1].value:
                    seventh726CurrentRow = seventh726CurrentRow + 1
                    write_to_file(seventh726Sheet, seventh726CurrentRow, row)

                elif "8th" in dataFileWS[cell1].value and "Service Elevator" in dataFileWS[cell1].value:
                    eigthSeCurrentRow = eigthSeCurrentRow + 1
                    write_to_file(eigthSeSheet, eigthSeCurrentRow, row)

                elif "9th" in dataFileWS[cell1].value and "East" in dataFileWS[cell1].value:
                    roofEastCurrentRow = roofEastCurrentRow + 1
                    write_to_file(roofEastSheet, roofEastCurrentRow, row)

                elif "9th" in dataFileWS[cell1].value and "West" in dataFileWS[cell1].value:
                    roofWestCurrentRow = roofWestCurrentRow + 1
                    write_to_file(roofWestSheet, roofWestCurrentRow, row)

                else:
                    restCurrentRow = restCurrentRow + 1

                    restSheet.cell(restCurrentRow, 1).value = dataFileWS[("{}{}".format("B", row))].value
                    restSheet.cell(restCurrentRow, 2).value = dataFileWS[("{}{}".format("C", row))].value
                    restSheet.cell(restCurrentRow, 3).value = dataFileWS[("{}{}".format("D", row))].value
                    restSheet.cell(restCurrentRow, 4).value = dataFileWS[("{}{}".format("F", row))].value
                    restSheet.cell(restCurrentRow, 5).value = dataFileWS[("{}{}".format("G", row))].value
                    restSheet.cell(restCurrentRow, 6).value = dataFileWS[("{}{}".format("H", row))].value
                    restSheet.cell(restCurrentRow, 7).value = dataFileWS[("{}{}".format("I", row))].value


    dataFileWB = getFile(path)
    dataFileWS = dataFileWB["Sheet1"]

    file = shutil.copy("Air Sample Results Trending Template.xlsx", savePath + "\\" + "AirSamplesCharts.xlsx")
    wb = openpyxl.load_workbook(file)

    #sheets
    sbSheet = wb["SB Svc Ele"]
    cwSheet = wb["CW Svc Ele"]
    bSheet = wb["B Svc Ele"]
    firstSeSheet = wb["1st Svc Ele"]
    secondSeSheet = wb["2nd Svc Ele"]
    secondHwSheet = wb["2nd Hall"]
    secondOutSheet = wb["2nd Out"]
    thirdSeSheet = wb["3rd Svc Ele"]
    fourthSeSheet = wb["4th Svc Ele"]
    fifthSeSheet = wb["5th Svc Ele"]
    sixthSeSheet = wb["6th Svc Ele"]
    seventhSeSheet = wb["7th Svc Ele"]
    seventh726Sheet = wb["7th Rm 726"]
    eigthSeSheet = wb["8th Svc Ele"]
    roofEastSheet = wb["9th East"]
    roofWestSheet = wb["9th West"]
    restSheet = wb["Rest"]

    find_location(dataFileWS)

    wb.close()
    wb.save(file)
