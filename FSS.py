import openpyxl
import os
import sys
import xlsxwriter
import statistics
import re

def main(path, savePath):
    grossTotalCounts = list()
    backgroundCounts = list()
    locationNumbers = list()
    locations = list()
    removableCounts = list()
    invalidSheets = list()
    netCPMRem = list()
    netActRem = list()
    netActTotal = list()
    netCPMTotal = list()
    invalidFiles = list()
    allNetAct = list()
    allRemAct = list()
    MDC = list()
    NDALocattion = list()
    ductSize = list()
    ductLength = list()
    NDAgrossCounts = list()
    NDAbackground = list()
    NDAefficiencyFactor= list()
    NDAnet = list()



    # Create the output folder
    if not os.path.isdir(savePath):
        os.mkdir(savePath)

    print(path)

    # create excel QC file
    QCworkbook = xlsxwriter.Workbook(savePath + '\\' + 'FSS_Trending.xlsx')
    QCworksheet = QCworkbook.add_worksheet("Data")
    StatisticSheet = QCworkbook.add_worksheet("Stats")

    # create columns with headers
    QCworksheet.write(0, 0, 'File Name')
    QCworksheet.write(0, 1, 'Sheet Name')
    QCworksheet.write(0, 2, 'Survey Number')
    QCworksheet.write(0, 3, 'Date')
    QCworksheet.write(0, 4, 'Survey Tech')
    QCworksheet.write(0, 5, 'Count room Tech')
    QCworksheet.write(0, 6, 'Date of Count Room')
    QCworksheet.write(0, 7, 'Survey Unit')
    QCworksheet.write(0, 8, 'Item Surveyed')
    QCworksheet.write(0, 9, 'No.')
    QCworksheet.write(0, 10, 'Description/Location')
    QCworksheet.write(0, 11, 'Active area of probe')
    QCworksheet.write(0, 12, 'Total Activity Instrument Efficiency')
    QCworksheet.write(0, 13, 'Total Activity Surface Efficiency')
    QCworksheet.write(0, 14, 'Gross counts of Total Activity')
    QCworksheet.write(0, 15, 'Background counts of Total activity')
    QCworksheet.write(0, 16, 'Total Background Count Time')
    QCworksheet.write(0, 17, 'Total Sample Count Time')
    QCworksheet.write(0, 18, 'MDC of Total activity')
    QCworksheet.write(0, 19, 'Net Activity of Total Activity')
    QCworksheet.write(0, 20, 'Removable Instrument Efficiency')
    QCworksheet.write(0, 21, 'Removable Instrument Surface Efficiency')
    QCworksheet.write(0, 22, 'Gross Counts of Removable')
    QCworksheet.write(0, 23, 'Background Counts of Removable')
    QCworksheet.write(0, 24, 'Removable Background Count Time (min')
    QCworksheet.write(0, 25, 'Removable Sample Count Time')
    QCworksheet.write(0, 26, 'MDC of Removable')
    QCworksheet.write(0, 27, 'Net Activity of Removable')

    # Create statistics sheet Headers
    StatisticSheet.write(0, 0, 'File Name')
    StatisticSheet.write(0, 1, 'Sheet Name')
    StatisticSheet.write(0, 2, 'Survey Number')
    StatisticSheet.write(0, 3, 'Total Activity Min')
    StatisticSheet.write(0, 4, 'Total Activity Max')
    StatisticSheet.write(0, 5, 'Total Activity Average')
    StatisticSheet.write(0, 6, 'Total Activity Standard Deviation')
    StatisticSheet.write(0, 7, 'Removable Min')
    StatisticSheet.write(0, 8, 'Removable Max')
    StatisticSheet.write(0, 9, 'Removable Average')
    StatisticSheet.write(0, 10, 'Removable Standard Deviation')

    StatisticSheet.write(0, 13, 'Overall Total Activity  Minimum')
    StatisticSheet.write(0, 14, 'Overall Total Activity Maximum')
    StatisticSheet.write(0, 15, 'Overall Total Activity Average')
    StatisticSheet.write(0, 16, 'Overall Total Activity Standard Deviation')
    StatisticSheet.write(0, 17, 'Overall Removable Minimum')
    StatisticSheet.write(0, 18, 'Overall Removable Maximum')
    StatisticSheet.write(0, 19, 'Overall Removable Average')
    StatisticSheet.write(0, 20, 'Overall Removable Standard Deviation')

    def resource_path(relative_path):
        base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.dirname(__file__)))
        return os.path.join(base_path, relative_path)

    def getListOfFiles(dirName):
        listOfFile = os.listdir(dirName)
        allFiles = list()

        for file in listOfFile:
            fullPath = os.path.join(dirName, file)
            if os.path.isdir(fullPath):
                allFiles = allFiles + getListOfFiles(fullPath)
            else:
                allFiles.append(fullPath)

        print(allFiles)

        return allFiles

    def find_cell(currentSheet, parameterToFind):
        for row in range(1, 40):
            for column in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":  # Here you can add or reduce the columns

                cell = "{}{}".format(column, row)

                if currentSheet[cell].value == parameterToFind:
                    print("the row is {0} and the column {1}".format(row, column))

                    print(currentSheet[cell].value)
                    print(cell)

                    return [row, column, currentSheet[cell]]

        return [0, 0, None]

    # This is used when we need to find a value and the searched term occurs more than once
    def find_date_cell(currentSheet, parameterToFind, occurenceNeeded):
        found = 0
        for row in range(1, 30):
            for column in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":

                modelVal = currentSheet[column + str(row)].value
                if modelVal != parameterToFind:
                    continue

                found += 1
                if found == occurenceNeeded:
                    return [row, column]

        return [None, None]

    def find_title_vals(currentSheet):
        # find Survey number
        sampleIdTitleCell = find_cell(currentSheet, "Survey No")
        if sampleIdTitleCell[2] is None:
            sampleIdTitleCell = find_cell(currentSheet, "Survey Number")
        sampleIdCell = find_title_data(currentSheet, sampleIdTitleCell)

        # find survey techs
        surveyTechTitleCell = find_cell(currentSheet, "Survey Tech")
        surveyTechCell = find_title_data(currentSheet, surveyTechTitleCell)

        # find Count room tech
        countRoomTechTitleCell = find_cell(currentSheet, "Count Room Tech")
        countRoomTechCell = find_title_data(currentSheet, countRoomTechTitleCell)

        # find type
        typeTitleCell = find_cell(currentSheet, "Survey Unit")
        typeCell = find_title_data(currentSheet, typeTitleCell)

        # find Item Surveyed
        locationTitleCell = find_cell(currentSheet, "Item Surveyed")
        if locationTitleCell[2] is None:
            locationTitleCell = find_cell(currentSheet, "Survey Number")
        locationCell = find_title_data(currentSheet, locationTitleCell)

        return [sampleIdCell, surveyTechCell, countRoomTechCell, typeCell, locationCell]

    def find_title_data(currentSheet, titleCell):
        row = int(titleCell[0])
        col = chr(ord(titleCell[1]) + 1)

        while type(currentSheet[col + str(row)]).__name__ == 'MergedCell':
            col = chr(ord(col) + 1)

        valueCell = currentSheet[col + str(row)]

        return valueCell

    def check_for_BettaGamma(num):
        found = 0
        for row in range(1, 30):
            for column in "GHIJKLMNOPQRSTUVWXYZ":

                modelVal = currentSheet[column + str(row)].value
                if modelVal != "Beta-Gamma":
                    continue

                found += 1
                if found == num:
                    return [row, column]

        return [None, None]

    def checkForMap():
        for row in range(2, 11):
            for column in "ABCDEFGHIJ":
                modelCell = "{}{}".format(column, row)
                if currentSheet[modelCell].value is not None:
                    return False

        return True

    def removePercent(var):
        # Find the %
        j = var.find('%')
        k = var[:j]

        # remove the percent from the string
        var = var.replace((str(k)+"%"), "")
        k = "." + k.replace(".", "")
        test20 = k + var

        return test20


    files = getListOfFiles(path)

    QCfileRow = 1
    SecondSheetRow = 1
    # Create Date format to write to xlsx files
    dateFormat = QCworkbook.add_format({'num_format': 'mm/dd/yyyy'})
    percentFormat = QCworkbook.add_format({'num_format': '0.00%'})

    for file in files:

        # For Openpyxl
        theFile = openpyxl.load_workbook(file)
        allSheetNames = theFile.sheetnames
        print(file)

        for x in allSheetNames:
            print("Current sheet name is {}".format(x))
            currentSheet = theFile[x]

            # If it is a map sheet skip
            currentSheetString = str(currentSheet)
            currentSheetString = currentSheetString[12:]
            checkBlankString = currentSheetString[0:5]
            currentSheetString = currentSheetString[0:3]
            if currentSheetString == "Map" or checkBlankString == "Blank":
                continue

            # Check which format it is
            betaRow, betaCol = check_for_BettaGamma(3)
            index = 0
            needToContinue = False

            # Prevent repeats be added to invalid Sheets list
            if betaRow is None or betaCol is None:
                needToContinue = True
                if len(invalidSheets) > 0:
                    if invalidSheets.count(file) == 0:
                        invalidSheets.append(file)

                else:
                    invalidSheets.append(file)

            # Find MDC Value and DPM for total activity as well as Removable DPM
            else:
                backgroundCounts.clear()
                grossTotalCounts.clear()
                removableCounts.clear()
                locationNumbers.clear()
                locations.clear()

                removableBetaRow, removableBetaCol = check_for_BettaGamma(4)
                countsCol = betaCol
                backgroundCol = chr(ord(betaCol) + 1)
                removableCountsCol = removableBetaCol

                n = 1
                # Go until it is not None
                while currentSheet[betaCol + str(betaRow + n)].value is None:
                    n += 1

                # There will always be at most 20 counts per survey
                for cell in range(n + 1, n + 21):
                    cellValue = currentSheet[countsCol + str(betaRow + cell)].value
                    backgroundValue = currentSheet[backgroundCol + str(betaRow + cell)].value
                    removableValue = currentSheet[removableCountsCol + str(removableBetaRow + cell)].value
                    locationNumber = currentSheet["A" + str(betaRow + cell)].value
                    location = currentSheet["B" + str(betaRow + cell)].value

                    # If there is nothing in both removable and total activity
                    if cellValue is None and removableValue is None:
                        continue

                    # If there is removable but no total activity
                    elif cellValue is None and removableValue is not None:
                        locationNumbers.append(locationNumber)
                        locations.append(location)
                        backgroundCounts.append(None)
                        grossTotalCounts.append(None)
                        removableCounts.append(removableValue)

                    # If there is total activity but no removable
                    elif cellValue is not None and removableValue is None:
                        locationNumbers.append(locationNumber)
                        locations.append(location)
                        backgroundCounts.append(backgroundValue)
                        grossTotalCounts.append(cellValue)
                        removableCounts.append(None)

                    # If there is both removable and total activity
                    else:
                        locationNumbers.append(locationNumber)
                        locations.append(location)
                        backgroundCounts.append(backgroundValue)
                        grossTotalCounts.append(cellValue)
                        removableCounts.append(removableValue)

                    index += 1

            if needToContinue is True:
                continue

            # Find totalEfficiency
            efficiencyRow, efficiencyCol = check_for_BettaGamma(1)
            efficiencyRow += 4
            totalEfficiency = currentSheet[efficiencyCol + str(efficiencyRow)].value
            if isinstance(totalEfficiency, str):
                try:
                    totalEfficiency = eval(removePercent(totalEfficiency[1:]))
                except:
                    continue
            ttemp, efficiencyCol = check_for_BettaGamma(2)
            remEfficiency = currentSheet[efficiencyCol + str(efficiencyRow)].value
            remSurfaceEfficiency = currentSheet[efficiencyCol + str(efficiencyRow + 1)].value

            # Find Removable Background
            efficiencyRow, efficiencyCol = check_for_BettaGamma(2)
            bkgRem = currentSheet[efficiencyCol + str(efficiencyRow + 9)].value
            remBkgCount = currentSheet[efficiencyCol + str(efficiencyRow + 7)].value
            remSampCount = currentSheet[efficiencyCol + str(efficiencyRow + 8)].value

            # Find Surface Efficiency
            efficiencyRow, efficiencyCol = check_for_BettaGamma(1)
            efficiencyRow += 5
            surfaceEfficiency = currentSheet[efficiencyCol + str(efficiencyRow)].value

            # Find Correction Factor
            correctionFactor = currentSheet[efficiencyCol + str(efficiencyRow + 1)].value
            totBkgCount = currentSheet[efficiencyCol + str(efficiencyRow + 2)].value
            totSampCount = currentSheet[efficiencyCol + str(efficiencyRow + 3)].value


            # ***********Find Title Data**********
            titleVals = find_title_vals(currentSheet)

            # find date
            dateTitleCell = find_cell(currentSheet, "Date")
            if dateTitleCell[2] is None or dateTitleCell[1] == 0:
                dateTitleCell = find_cell(currentSheet, "Date Counted")
            dateCell = find_title_data(currentSheet, dateTitleCell)

            # Find Count Room Date Counted
            dateTitleCell = find_date_cell(currentSheet, "Date Counted", 2)
            if dateTitleCell[1] is None or dateTitleCell[1] == 0:
                dateTitleCell = find_date_cell(currentSheet, "Date Counted", 1)
            secondDateCell = find_title_data(currentSheet, dateTitleCell)

            # Find the Name of the worksheet
            currentSheetString = str(currentSheet)
            currentSheetString = currentSheetString[12:]
            currentSheetString = currentSheetString[:-2]

            netCPMRem.clear()
            netActRem.clear()
            netActTotal.clear()
            netCPMTotal.clear()
            MDC.clear()

            # Find DPMs
            badFile = False
            for i in range(0, len(removableCounts)):

                if removableCounts[i] is None:
                    netCPMRem.append(None)
                    netActRem.append(None)

                elif type(bkgRem) != str:
                    netCPMRem.append(removableCounts[i] - (bkgRem / 60))
                    netActRem.append(netCPMRem[i] / remEfficiency)

                else:
                    try:
                        bkgRem = bkgRem[1:]
                        bkgRem = eval(bkgRem)

                        # Better security but we need to test much more
                        # bkgRem = int(sympy.sympify(bkgRem))

                        netCPMRem.append(removableCounts[i] - (bkgRem / 60))
                        netActRem.append(netCPMRem[i] / remEfficiency)

                    except:
                        if invalidFiles.count(file) == 0:
                            invalidFiles.append(file)
                            badFile = True

            # total activity calculations
            for i in range(0, len(grossTotalCounts)):
                if grossTotalCounts[i] is None:
                    netCPMTotal.append(None)
                    netActTotal.append(None)

                elif type(bkgRem) != str:
                    netCPMTotal.append(grossTotalCounts[i] - (backgroundCounts[i]))
                    netActTotal.append((netCPMTotal[i] / (totalEfficiency * surfaceEfficiency)) * correctionFactor)
                    MDC.append(((3 + 3.29 * ((backgroundCounts[i]) * 2) ** 0.5) / (totalEfficiency * surfaceEfficiency)) * correctionFactor)

                else:
                    try:
                        backgroundCounts[i] = backgroundCounts[i][1:]
                        backgroundCounts[i] = eval(backgroundCounts[i])

                        # Better security but we need to test much more
                        # backgroundCounts[i] = int(sympy.sympify(backgroundCounts[i]))

                        netCPMTotal.append(grossTotalCounts[i] - (backgroundCounts[i]))
                        netActTotal.append((netCPMTotal[i] / (totalEfficiency * surfaceEfficiency)) * correctionFactor)
                        MDC.append(((3 + 3.29 * ((backgroundCounts[i]) * 1 * (1 + (1 / 1))) ** 0.5) / (totalEfficiency * surfaceEfficiency)) * 1.25)

                    except:
                        if invalidFiles.count(file) == 0:
                            invalidFiles.append(file)
                            badFile = True

            # Write the results to the QC file
            # Write the current Worksheet
            area = re.sub('\^2', '²', "100 cm^2")
            try:
                bkgMDC = (3 + 3.29 * (((bkgRem / 60) * 1 * (1 + (1 / 60))) ** 0.5)) / (remEfficiency)
            except:
                print("No removable efficiency")

            head, tail = os.path.split(file)
            if badFile is True:
                continue
            for i in range(0, len(removableCounts)):
                QCworksheet.write(QCfileRow, 0, tail)  # File Name
                QCworksheet.write(QCfileRow, 1, currentSheetString) # Sheet name
                QCworksheet.write(QCfileRow, 2, titleVals[0].value)  # Survey Number
                QCworksheet.write(QCfileRow, 3, dateCell.value, dateFormat)  # Date
                QCworksheet.write(QCfileRow, 4, titleVals[1].value)  # Survey Tech
                QCworksheet.write(QCfileRow, 5, titleVals[2].value)  # Count room Tech
                QCworksheet.write(QCfileRow, 6, secondDateCell.value, dateFormat)  # Date of Count Room
                QCworksheet.write(QCfileRow, 7, titleVals[3].value)  # Survey Unit
                QCworksheet.write(QCfileRow, 8, titleVals[4].value)  # Item Surveyed
                QCworksheet.write(QCfileRow, 9, locationNumbers[i])
                QCworksheet.write(QCfileRow, 10, locations[i])
                QCworksheet.write(QCfileRow, 11, area)  # Active probe area
                QCworksheet.write(QCfileRow, 12, totalEfficiency, percentFormat)  # totalEfficiency
                QCworksheet.write(QCfileRow, 13, surfaceEfficiency, percentFormat)  # Total Activity Surface Efficiency
                QCworksheet.write(QCfileRow, 14, grossTotalCounts[i])  # Gross Counts Total
                QCworksheet.write(QCfileRow, 15, backgroundCounts[i])  # Background total activity
                QCworksheet.write(QCfileRow, 16, totBkgCount)  # Total background count time
                QCworksheet.write(QCfileRow, 17, totSampCount)  # Total sample count time
                QCworksheet.write(QCfileRow, 18, round(MDC[i]))  # MDC total activity
                QCworksheet.write(QCfileRow, 19, round(netActTotal[i]))  # DPM total activity
                QCworksheet.write(QCfileRow, 20, remEfficiency, percentFormat)  # Removable Instrument Efficiency
                QCworksheet.write(QCfileRow, 21, remSurfaceEfficiency, percentFormat)  # Removable Surface Efficiency
                QCworksheet.write(QCfileRow, 22, removableCounts[i])  # Gross removable Counts
                QCworksheet.write(QCfileRow, 23, round(bkgRem))  # Removable activity background
                QCworksheet.write(QCfileRow, 24, remBkgCount)  # Removable background count time
                QCworksheet.write(QCfileRow, 25, remSampCount)  # Removable sample count time
                QCworksheet.write(QCfileRow, 26, round(bkgMDC))  # Removable MDC
                QCworksheet.write(QCfileRow, 27, round(netActRem[i]))  # Removable DPM

                QCfileRow += 1

            allNetAct.extend(netActTotal)
            allRemAct.extend(netActRem)

            # Find the statistics
            netActTotal = list(filter(None, netActTotal))
            netActRem = list(filter(None, netActRem))

            if len(netActTotal) != 0:
                totalAverage = sum(netActTotal) / len(netActTotal)
                totalMax = max(netActTotal)
                totalMin = min(netActTotal)
                totalStdDev = statistics.pstdev(netActTotal)
            else:
                totalAverage = None
                totalMax = None
                totalMin = None
                totalStdDev = None

            if len(netActRem) != 0:
                removableAvg = sum(netActRem) / len(netActRem)
                removableMax = max(netActRem)
                removableMin = min(netActRem)
                removableStdDev = statistics.pstdev(netActRem)
            else:
                removableAvg = None
                removableMax = None
                removableMin = None
                removableStdDev = None

            StatisticSheet.write(SecondSheetRow, 0, tail)  # File Name
            StatisticSheet.write(SecondSheetRow, 1, currentSheetString)
            StatisticSheet.write(SecondSheetRow, 2, titleVals[0].value)  # Survey Number
            StatisticSheet.write(SecondSheetRow, 3, totalMin)
            StatisticSheet.write(SecondSheetRow, 4, totalMax)
            StatisticSheet.write(SecondSheetRow, 5, totalAverage)
            StatisticSheet.write(SecondSheetRow, 6, totalStdDev)
            StatisticSheet.write(SecondSheetRow, 7, removableMin)
            StatisticSheet.write(SecondSheetRow, 8, removableMax)
            StatisticSheet.write(SecondSheetRow, 9, removableAvg)
            StatisticSheet.write(SecondSheetRow, 10, removableStdDev)

            SecondSheetRow += 1

        theFile.close()

    """redo but for the other list"""
    print("IN THE SCAN SHEET")
    del grossTotalCounts
    del backgroundCounts
    del removableCounts
    del locationNumbers
    del locations
    grossTotalCounts = list()
    removableCounts = list()
    locationNumbers = list()
    locations = list()

    for file in invalidSheets:
        theFile = openpyxl.load_workbook(file)
        allSheetNames = theFile.sheetnames
        print(file)
        print("All sheet names {} ".format(theFile.sheetnames))

        for x in allSheetNames:
            print("Current sheet name is {}".format(x))
            currentSheet = theFile[x]

            # If it is a map sheet skip
            currentSheetString = str(currentSheet)
            currentSheetString = currentSheetString[12:]
            checkBlankString = currentSheetString[0:5]
            currentSheetString = currentSheetString[0:3]
            if currentSheetString == "Map" or checkBlankString == "Blank":
                continue

            betaRow, betaCol = check_for_BettaGamma(3)
            if betaRow is not None or betaCol is not None:
                continue

            if checkForMap():
                continue

            betaRow, betaCol = check_for_BettaGamma(1)
            if betaRow is None or betaCol is None:
                continue

            else:
                grossTotalCounts.clear()

                betaRow, betaCol = check_for_BettaGamma(2)

                n = 1
                # Go until it is not None
                while currentSheet[betaCol + str(betaRow + n)].value != "Ambient Bkg":
                    print("currentsheet")
                    print(currentSheet[betaCol + str(betaRow + n)].value)
                    n += 1

                # There will always be at most 20 counts per survey
                for cell in range(n + 1, n + 21):
                    cellValue = currentSheet[betaCol + str(betaRow + cell)].value
                    locationNumber = currentSheet["A" + str(betaRow + cell)].value
                    location = currentSheet["B" + str(betaRow + cell)].value

                    # If there is nothing in total activity
                    if cellValue is None:
                        continue

                    # If there is total activity
                    else:
                        grossTotalCounts.append(cellValue)
                        locationNumbers.append(locationNumber)
                        locations.append(location)

                    index += 1

            # Find efficiency
            efficiencyRow, efficiencyCol = check_for_BettaGamma(1)
            efficiencyRow += 4
            totalEfficiency = currentSheet[efficiencyCol + str(efficiencyRow)].value

            # Find Background
            efficiencyRow, efficiencyCol = check_for_BettaGamma(1)
            efficiencyRow += 5
            bkgTotal = currentSheet[efficiencyCol + str(efficiencyRow)].value

            # ***********Find Title Data**********
            titleVals = find_title_vals(currentSheet)
            print("After titlevals")

            # find date
            dateTitleCell = find_cell(currentSheet, "Date")
            if dateTitleCell[2] is None or dateTitleCell[1] == 0:
                dateTitleCell = find_cell(currentSheet, "Date Counted")
            dateCell = find_title_data(currentSheet, dateTitleCell)
            print("After date")

            # Find Count Room Date Counted
            dateTitleCell = find_date_cell(currentSheet, "Date Counted", 2)
            if dateTitleCell[1] is None or dateTitleCell[1] == 0:
                dateTitleCell = find_date_cell(currentSheet, "Date Counted", 1)
            secondDateCell = find_title_data(currentSheet, dateTitleCell)
            print("After second date")

            # Find Surface Efficiency
            efficiencyRow, efficiencyCol = check_for_BettaGamma(1)
            efficiencyRow += 5
            surfaceEfficiency = currentSheet[efficiencyCol + str(efficiencyRow)].value

            # Find Correction Factor
            efficiencyRow, efficiencyCol = check_for_BettaGamma(1)
            efficiencyRow += 6
            correctionFactor = currentSheet[efficiencyCol + str(efficiencyRow)].value

            currentSheetString = str(currentSheet)
            currentSheetString = currentSheetString[12:]
            currentSheetString = currentSheetString[:-2]

            netActTotal.clear()
            netCPMTotal.clear()

            # Find DPMs
            for i in range(0, len(grossTotalCounts)):
                if grossTotalCounts[i] is None:
                    netCPMTotal.append(None)
                    netActTotal.append(None)
                else:
                    netCPMTotal.append(grossTotalCounts[i] - bkgTotal)
                    netActTotal.append((netCPMTotal[i] / (totalEfficiency * surfaceEfficiency)) * correctionFactor)

            print("Adding data to file.")
            head, tail = os.path.split(file)
            for i in range(0, len(grossTotalCounts)):
                QCworksheet.write(QCfileRow, 0, tail)  # File Name
                QCworksheet.write(QCfileRow, 1, currentSheetString)
                QCworksheet.write(QCfileRow, 2, titleVals[0].value)  # Survey Number
                QCworksheet.write(QCfileRow, 3, dateCell.value, dateFormat)  # Date
                QCworksheet.write(QCfileRow, 4, titleVals[1].value)  # Survey Tech
                QCworksheet.write(QCfileRow, 5, titleVals[2].value)  # Count room Tech
                QCworksheet.write(QCfileRow, 6, secondDateCell.value, dateFormat)  # Date of Count Room
                QCworksheet.write(QCfileRow, 7, titleVals[3].value)  # Survey Unit
                QCworksheet.write(QCfileRow, 8, titleVals[4].value)  # Item Surveyed
                QCworksheet.write(QCfileRow, 9, locationNumbers[i])
                QCworksheet.write(QCfileRow, 10, locations[i])
                QCworksheet.write(QCfileRow, 11, area)
                QCworksheet.write(QCfileRow, 12, totalEfficiency, percentFormat)  # totalEfficiency
                QCworksheet.write(QCfileRow, 13, surfaceEfficiency, percentFormat)
                QCworksheet.write(QCfileRow, 14, grossTotalCounts[i])  # Gross Counts Total
                QCworksheet.write(QCfileRow, 15, netActTotal[i])  # DPM total activity
                QCworksheet.write(QCfileRow, 16, totBkgCount)  # Total background count time
                QCworksheet.write(QCfileRow, 17, totSampCount)  # Total sample count time

                QCfileRow += 1

            # Find the statistics
            allNetAct.extend(netActTotal)

            # Find the statistics
            netActTotal = list(filter(None, netActTotal))

            if len(netActTotal) != 0:
                totalAverage = sum(netActTotal) / len(netActTotal)
                totalMax = max(netActTotal)
                totalMin = min(netActTotal)
                totalStdDev = statistics.pstdev(netActTotal)
            else:
                totalAverage = None
                totalMax = None
                totalMin = None
                totalStdDev = None

            StatisticSheet.write(SecondSheetRow, 0, tail)  # File Name
            StatisticSheet.write(SecondSheetRow, 1, currentSheetString)
            StatisticSheet.write(SecondSheetRow, 2, titleVals[0].value)  # Survey Number
            StatisticSheet.write(SecondSheetRow, 3, currentSheetString)  # Current Sheet
            StatisticSheet.write(SecondSheetRow, 4, totalMin)
            StatisticSheet.write(SecondSheetRow, 5, totalMax)
            StatisticSheet.write(SecondSheetRow, 6, totalAverage)
            StatisticSheet.write(SecondSheetRow, 7, totalStdDev)

            SecondSheetRow += 1

            grossTotalCounts.clear()
            removableCounts.clear()
            locationNumbers.clear()
            locations.clear()

        theFile.close()
        theFile.save(file)
        grossTotalCounts.clear()

    # Check for NDA sheets
    NDASheetRow = 1
    for file in files:

        # For Openpyxl
        theFile = openpyxl.load_workbook(file)
        allSheetNames = theFile.sheetnames
        print(file)

        for x in allSheetNames:
            print("Current sheet name is {}".format(x))
            currentSheet = theFile[x]

            # If it is a map sheet skip
            currentSheetString = str(currentSheet)
            currentSheetString = currentSheetString[12:]
            currentSheetName = currentSheetString[:-2]
            checkNDAString = currentSheetString[0:3]

            if checkNDAString == "NDA":

                # Start Setup
                try:
                    NDASheet = QCworkbook.add_worksheet("NDA")

                    NDASheet.write(0, 0, 'File Name')
                    NDASheet.write(0, 1, 'Sheet Name')
                    NDASheet.write(0, 2, 'Survey Number')
                    NDASheet.write(0, 3, 'Date')
                    NDASheet.write(0, 4, 'Survey Tech')
                    NDASheet.write(0, 5, 'Survey Unit')
                    NDASheet.write(0, 6, 'Item Surveyed')
                    NDASheet.write(0, 7, 'Active area of probe')
                    NDASheet.write(0, 8, 'Instrument Model')
                    NDASheet.write(0, 9, 'Instrument SN')
                    NDASheet.write(0, 10, 'Instrument Cal Due Date')
                    NDASheet.write(0, 11, 'Count Time (min)')
                    NDASheet.write(0, 12, 'Description/Location')
                    NDASheet.write(0, 13, 'Duct Size')
                    NDASheet.write(0, 14, 'Duct Length')
                    NDASheet.write(0, 15, 'Gross Counts')
                    NDASheet.write(0, 16, 'Background Counts')
                    NDASheet.write(0, 17, 'Efficiency Factor')
                    NDASheet.write(0, 18, 'Net Activity')

                except:
                    print("NDA sheet already made")

                if checkForMap():
                    continue

                surveyNumber = currentSheet["D1"].value
                dateSurveyed = currentSheet["D2"].value
                techs = currentSheet["D3"].value
                surveyUnit = currentSheet["D4"].value
                instrumentModel = currentSheet["W8"].value
                SN = currentSheet["W9"].value
                calDueDate = currentSheet["W10"].value
                itemSurveyed = currentSheet["J1"].value

                NDAefficiencyFactor.clear()
                NDALocattion.clear()
                NDAbackground.clear()
                NDAgrossCounts.clear()
                NDAnet.clear()
                ductSize.clear()
                ductLength.clear()


                # Get Static Values
                locationNumbers.clear()
                for i in range(1, 21):
                    if currentSheet["B" + str(i + 20)].value is not None:
                        locationNumbers.append(currentSheet["A" + str(i + 20)].value)
                        NDALocattion.append(currentSheet["B" + str(i + 20)].value)
                        ductSize.append(currentSheet["N" + str(i + 20)].value)
                        ductLength.append(currentSheet["P" + str(i + 20)].value)
                        NDAgrossCounts.append(currentSheet["R" + str(i + 20)].value)
                        NDAbackground.append(currentSheet["T" + str(i + 20)].value)
                        NDAefficiencyFactor.append(currentSheet["V" + str(i + 20)].value)

                for i in range(0, len(NDALocattion)):
                    NDAnet.append((NDAgrossCounts[i] - NDAbackground[i]) * NDAefficiencyFactor[i])

                head, tail = os.path.split(file)
                for i in range(0, len(NDALocattion)):
                    NDASheet.write(NDASheetRow, 0, tail)  # File Name
                    NDASheet.write(NDASheetRow, 1, currentSheetName)  # Sheet name
                    NDASheet.write(NDASheetRow, 2, surveyNumber)  # Survey Number
                    NDASheet.write(NDASheetRow, 3, dateSurveyed, dateFormat)  # Date
                    NDASheet.write(NDASheetRow, 4, techs)  # Survey Tech
                    NDASheet.write(NDASheetRow, 5, surveyUnit)  # Survey Unit
                    NDASheet.write(NDASheetRow, 6, itemSurveyed)
                    NDASheet.write(NDASheetRow, 7, 1)
                    NDASheet.write(NDASheetRow, 8, instrumentModel)
                    NDASheet.write(NDASheetRow, 9, SN)
                    NDASheet.write(NDASheetRow, 10, calDueDate, dateFormat)
                    NDASheet.write(NDASheetRow, 11, 1)
                    NDASheet.write(NDASheetRow, 12, locationNumbers[i])
                    NDASheet.write(NDASheetRow, 13, NDALocattion[i])
                    NDASheet.write(NDASheetRow, 14, ductSize[i])
                    NDASheet.write(NDASheetRow, 15, ductLength[i])
                    NDASheet.write(NDASheetRow, 16, NDAgrossCounts[i])
                    NDASheet.write(NDASheetRow, 17, NDAbackground[i])
                    NDASheet.write(NDASheetRow, 18, NDAefficiencyFactor[i])
                    NDASheet.write(NDASheetRow, 19, round(NDAnet[i]))

                    NDASheetRow += 1

    # Find overall stats
    allNetAct = list(filter(None, allNetAct))
    allRemAct = list(filter(None, allRemAct))
    allTotalAverage = 0
    allTotalMax = 0
    allTotalMin = 0
    allTotalStdDev = 0
    allRemovableAvg = 0
    allRemovableMax = 0
    allRemovableMin = 0
    allRemovableStdDev = 0


    if len(allNetAct) != 0:
        allTotalAverage = sum(allNetAct) / len(allNetAct)
        allTotalMax = max(allNetAct)
        allTotalMin = min(allNetAct)
        allTotalStdDev = statistics.pstdev(allNetAct)

    if len(allRemAct) != 0:
        allRemovableAvg = sum(allRemAct) / len(allRemAct)
        allRemovableMax = max(allRemAct)
        allRemovableMin = min(allRemAct)
        allRemovableStdDev = statistics.pstdev(allRemAct)

    StatisticSheet.write(1, 13, allTotalMin)
    StatisticSheet.write(1, 14, allTotalMax)
    StatisticSheet.write(1, 15, allTotalAverage)
    StatisticSheet.write(1, 16, allTotalStdDev)
    StatisticSheet.write(1, 17, allRemovableMin)
    StatisticSheet.write(1, 18, allRemovableMax)
    StatisticSheet.write(1, 19, allRemovableAvg)
    StatisticSheet.write(1, 20, allRemovableStdDev)

    if len(invalidFiles) > 0:
        FailedSheet = QCworkbook.add_worksheet()
        FailedSheet.write(0, 0, 'Invalid Files')
        x = 1
        for file in invalidFiles:
            FailedSheet.write(x, 0, file)

    QCworkbook.close()
    os.startfile(savePath + '\\' + 'FSS_Trending.xlsx')
