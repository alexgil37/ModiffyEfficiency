import openpyxl
import json
import os
import sys
import xlsxwriter
from pycel import ExcelCompiler

def main(path, savePath):
    # Create the output folder
    if os.path.isfile(savePath):
        os.mkdir(savePath)

    print(path)

    # create excel QC file
    QCworkbook = xlsxwriter.Workbook(savePath + '\\' + 'AirSamplesTrending.xlsx')
    QCworksheet = QCworkbook.add_worksheet()

    # create columns with headers
    QCworksheet.write(0, 0, 'File Name')
    QCworksheet.write(0, 1, 'Sample ID')
    QCworksheet.write(0, 2, 'Date')
    QCworksheet.write(0, 3, 'Location')
    QCworksheet.write(0, 4, 'Sample Type')
    QCworksheet.write(0, 5, 'Alpha Activity')
    QCworksheet.write(0, 6, 'Alpha MDC')
    QCworksheet.write(0, 7, 'Beta Activity')
    QCworksheet.write(0, 8, 'Beta MDC')


    def resource_path(relative_path):
        base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.dirname(__file__)))
        return os.path.join(base_path, relative_path)


    def getListOfFiles(dirName):
        listOfFile = os.listdir(dirName)
        allFiles = list()

        for file in listOfFile:
            fullPath = os.path.join(dirName, file)
            if os.path.isdir(fullPath):
                allFiles = allFiles + getListOfFiles(fullPath)
            else:
                allFiles.append(fullPath)

        print(allFiles)

        return allFiles


    def find_cell(currentSheet, parameterToFind):
        for row in range(1, 40):
            for column in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":  # Here you can add or reduce the columns

                cell = "{}{}".format(column, row)

                if currentSheet[cell].value == parameterToFind:
                    print("the row is {0} and the column {1}".format(row, column))

                    print(currentSheet[cell].value)
                    print(cell)

                    return [row, column, currentSheet[cell]]

        return [0, 0, None]

    def find_date(currentSheet, dateTitleCell):
        row = int(dateTitleCell[0]) + 1
        print(row)
        col = dateTitleCell[1]
        print(col)

        while type(currentSheet[col + str(row)]).__name__ == 'MergedCell':
            row = row + 1

        dateCell = currentSheet[col + str(row)]

        return dateCell

    def find_sampleId(currentSheet, sampleIdTitleCell):
        row = int(sampleIdTitleCell[0]) + 1
        print(row)
        col = sampleIdTitleCell[1]
        print(col)

        while type(currentSheet[col + str(row)]).__name__ == 'MergedCell':
            row = row + 1

        sampleIdCell = currentSheet[col + str(row)]

        return sampleIdCell

    def find_location(currentSheet, locationTitleCell):
        row = int(locationTitleCell[0]) + 1
        print(row)
        col = locationTitleCell[1]
        print(col)

        while type(currentSheet[col + str(row)]).__name__ == 'MergedCell':
            row = row + 1

        locationCell = currentSheet[col + str(row)]

        return locationCell

    def find_type(currentSheet, typeTytleCell):
        row = int(typeTytleCell[0]) + 1
        print(row)
        col = typeTytleCell[1]
        print(col)

        while type(currentSheet[col + str(row)]).__name__ == 'MergedCell':
            row = row + 1

        typeCell = currentSheet[col + str(row)]

        return typeCell

    #works for activity and mdc
    def find_activity(currentSheet, activityTitleCell):
        row = int(activityTitleCell[0])
        print(row)
        col = activityTitleCell[1]
        print(col)

        return [(chr(ord(col) + 4) + str(row)), (chr(ord(col) + 7) + str(row))]


    files = getListOfFiles(path)

    QCfileRow = 1

    # Create Date format to write to xlsx files
    dateFormat = QCworkbook.add_format({'num_format': 'mm/dd/yyyy'})

    for file in files:
        excel = ExcelCompiler(filename=file)
        theFile = openpyxl.load_workbook(file)
        allSheetNames = theFile.sheetnames

        print("All sheet names {} ".format(theFile.sheetnames))

        for x in allSheetNames:
            print("Current sheet name is {}" .format(x))
            currentSheet = theFile[x]

            #find date
            dateTitleCell = find_cell(currentSheet, "Date/Time On")
            dateCell = find_date(currentSheet, dateTitleCell)

            #find sample id
            sampleIdTitleCell = find_cell(currentSheet, "Sample ID")
            sampleIdCell = find_sampleId(currentSheet, sampleIdTitleCell)

            #find location
            locationTitleCell = find_cell(currentSheet, "Air Sample Location")
            locationCell = find_location(currentSheet, locationTitleCell)

            #find type
            typeTitleCell = find_cell(currentSheet, "Sample Type")
            typeCell = find_type(currentSheet, typeTitleCell)

            #find activity
            activityTitleCell = find_cell(currentSheet, "Activity (µCi/mL)")
            activityCell = find_activity(currentSheet, activityTitleCell)

            #find mdc
            mdcTitleCell = find_cell(currentSheet, "MDC (µCi/mL)")
            mdcCell = find_activity(currentSheet, mdcTitleCell)

            #find the file name
            head, tail = os.path.split(file)

            # Find the Name of the worksheet
            currentSheetString = str(currentSheet)
            currentSheetString = currentSheetString[12:]
            currentSheetString = currentSheetString[:-2]

            # Find the values of alpha and beta activity
            alphaCell = currentSheetString + "!" + str(activityCell[0])
            alphaActivity = excel.evaluate(alphaCell)

            betaCell = currentSheetString + "!" + str(activityCell[1])
            betaActivity = excel.evaluate(betaCell)

            # Find the values of alpha and beta mdc
            mdcalphaCell = currentSheetString + "!" + str(mdcCell[0])
            alphaMdc = excel.evaluate(mdcalphaCell)

            mdcbetaCell = currentSheetString + "!" + str(mdcCell[1])
            betaMdc = excel.evaluate(mdcbetaCell)

            # Write the results to the QC file
            # Write the current Worksheet
            head, tail = os.path.split(file)
            QCworksheet.write(QCfileRow, 0, tail)
            QCworksheet.write(QCfileRow, 1, sampleIdCell.value)
            QCworksheet.write(QCfileRow, 2, dateCell.value, dateFormat)
            QCworksheet.write(QCfileRow, 3, locationCell.value)
            QCworksheet.write(QCfileRow, 4, typeCell.value)
            QCworksheet.write(QCfileRow, 5, alphaActivity)
            QCworksheet.write(QCfileRow, 7, betaActivity)
            QCworksheet.write(QCfileRow, 6, alphaMdc)
            QCworksheet.write(QCfileRow, 8, betaMdc)

            QCfileRow += 1

        theFile.close()
        theFile.save(file)

    QCworkbook.close()
    os.startfile(savePath + '\\' + 'AirSamplesTrending.xlsx')
