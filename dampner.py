import openpyxl
import re
import os
import sys
import xlsxwriter

def main(path, savePath):
    # Create the output folder
    if not os.path.isdir(savePath):
        os.mkdir(savePath)

    # create excel QC file
    QCworkbook = xlsxwriter.Workbook(savePath + '\\' + 'Dampner.xlsx')
    QCworksheet = QCworkbook.add_worksheet()

    # create columns with headers
    QCworksheet.write(0, 0, 'File Name')

    def getListOfFiles(dirName):
        listOfFile = os.listdir(dirName)
        allFiles = list()

        for file in listOfFile:
            fullPath = os.path.join(dirName, file)
            if os.path.isdir(fullPath):
                allFiles = allFiles + getListOfFiles(fullPath)
            else:
                allFiles.append(fullPath)

        print(allFiles)

        return allFiles

    files = getListOfFiles(path)
    listOfFilesWithDampner = list()

    for file in files:
        try:
            # For Openpyxl
            theFile = openpyxl.load_workbook(file)
        except:
            continue
        allSheetNames = theFile.sheetnames
        print(file)
        print("All sheet names {} ".format(theFile.sheetnames))

        for x in allSheetNames:
            print("Current sheet name is {}".format(x))
            currentSheet = theFile[x]

            # If it is a map sheet skip
            currentSheetString = str(currentSheet)
            currentSheetString = currentSheetString[12:]
            checkBlankString = currentSheetString[0:5]
            currentSheetString = currentSheetString[0:3]
            if currentSheetString == "Map" or checkBlankString == "Blank":
                continue

            # Check each cell
            for row in range(1, 40):
                for column in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":  # Here you can add or reduce the columns
                    cell = "{}{}".format(column, row)
                    value = str(currentSheet[cell].value)
                    value =value

                    # Check if dampner is in in the cell and if it is already in the list
                    if "damper" in value or "dampener" in value or "dampner" in value:
                        if len(listOfFilesWithDampner) > 0:
                            if listOfFilesWithDampner.count(file) == 0:
                                head, tail = os.path.split(file)
                                listOfFilesWithDampner.append(tail)
                        else:
                            head, tail = os.path.split(file)
                            listOfFilesWithDampner.append(tail)

                        continue

    # print list
    i = 0
    for x in listOfFilesWithDampner:
        QCworksheet.write(i+1, 0, x)
        i += 1

    QCworkbook.close()
