import openpyxl
import json
import os
import xlsxwriter
from distutils.dir_util import copy_tree



def main(path, savePath):
    # Create the output folder
    if os.path.isfile(savePath):
        os.mkdir(savePath)

    # copy original files into designated filepath
    fileSavePath = os.path.join(savePath, "Output Files")
    if os.path.isfile(savePath):
        os.mkdir(fileSavePath)
    copy_tree(path, fileSavePath)

    # create excel QC file
    QCworkbook = xlsxwriter.Workbook(savePath + '\\' + 'QC.xlsx')
    QCworksheet = QCworkbook.add_worksheet()

    # create columns with headers
    QCworksheet.write(0, 0, 'File Name')
    QCworksheet.write(0, 1, 'Sheet Name')
    QCworksheet.write(0, 2, 'Instrument S/N')
    QCworksheet.write(0, 3, 'Cal DueDate')
    QCworksheet.write(0, 4, 'Old Efficiency')
    QCworksheet.write(0, 5, 'New Efficiency')



    def getListOfFiles(dirName):
        listOfFile = os.listdir(dirName)
        allFiles = list()

        for file in listOfFile:
            fullPath = os.path.join(dirName, file)
            if os.path.isdir(fullPath):
                allFiles = allFiles + getListOfFiles(fullPath)
            else:
                allFiles.append(fullPath)

        # print(allFiles)

        return allFiles


    def find_instrument_model_cell(currentSheet):
        for row in range(1, 50):
            for column in "ABCDEFGHIJKLMNOPQRSTUV":  # Here you can add or reduce the columns
                modelCell = "{}{}".format(column, row)
                if currentSheet[modelCell].value == instrumentModel:
                    print("the row is {0} and the column {1}" .format(row, column))
                    print(currentSheet[modelCell].value)
                    print(modelCell)

                    return [row, column, modelCell]

        return [0, 0, None]


    def find_instrument_sn_cell(instModelRow, instModelColumn):
        snRow = str(int(instModelRow) + 1)
        snCol = instModelColumn
        snCell = currentSheet[snCol + snRow]

        print('xxxxxxxxxxx')
        print(snRow)
        print(snCol)
        print(snCell.value)

        return snCell

    def find_cal_due_date(instModelRow, instModelColumn):
        calRow = str(int(instModelRow) + 2)
        calCol = chr(ord(instModelColumn))
        calCell = currentSheet[calCol + calRow]

        print(calCell.value)
        print(calRow)
        print(calCol)

        return calCell


    def find_instrument_efficiency(instModelRow, instModelColumn):
        effRow = str(int(instModelRow) + 3)
        effCol = chr(ord(instModelColumn) + 2)
        effCell = currentSheet[effCol + effRow]

        # In case it is 3 cells merged instead of 2
        if type(effCell).__name__ == 'MergedCell':
            effCol = chr(ord(instModelColumn) + 3)
            effCell = currentSheet[effCol + effRow]


        print(effCell.value)
        print(effRow)
        print(effCol)

        return effCell


    def modify_efficiency(instSNcell, instEfficiencyCell):
        for inst in instrumentsData:
            if inst['sn'] == instSNcell.value:
                instEfficiencyCell.value = inst['betaEfficiency']

                return [inst['sn'], inst['betaEfficiency']]

        return [None, None]



    instrumentModel = '2360/43-93'
    #instrumentId = '227413/PR295918'

    filesWithNoMatchingSN = list()
    sheetsOfFilesWithNoMatchingSN = list()

    files = getListOfFiles(fileSavePath)

    with open('package.json') as instruments_file:
        instrumentsData = json.load(instruments_file)

    QCfileRow = 1
    # Create Date format to write to xlsx files
    dateFormat = QCworkbook.add_format({'num_format': 'mm/dd/yyyy'})

    for file in files:
        theFile = openpyxl.load_workbook(file)
        allSheetNames = theFile.sheetnames

        print("All sheet names {} ".format(theFile.sheetnames))

        for x in allSheetNames:
            # print("Current sheet name is {}" .format(x))
            currentSheet = theFile[x]
            instModelRow = find_instrument_model_cell(currentSheet)[0]
            instModelColumn = find_instrument_model_cell(currentSheet)[1]
            instModelCell = find_instrument_model_cell(currentSheet)[2]

            print("The cell is {}, the row is {} and the column is {} ".format(instModelCell, instModelRow, instModelColumn))

            # Check if Model Number is blank
            if instModelCell is None:
                continue

            instSNcell = find_instrument_sn_cell(instModelRow, instModelColumn)
            instEfficiencyCell = find_instrument_efficiency(instModelRow, instModelColumn)
            oldinstEfficiencyCell = instEfficiencyCell.value
            instCalDueDate = find_cal_due_date(instModelRow, instModelColumn)
            serialNumber = modify_efficiency(instSNcell, instEfficiencyCell)

            # Finding all surveys that do not have a SN
            if serialNumber[0] is None:
                filesWithNoMatchingSN.append(file)
                sheetsOfFilesWithNoMatchingSN.append(currentSheet)

            # Write the results to the QC file
            QCworksheet.write(QCfileRow, 0, file)
            QCworksheet.write(QCfileRow, 1, str(currentSheet))
            QCworksheet.write(QCfileRow, 2, instSNcell.value)
            QCworksheet.write(QCfileRow, 3, instCalDueDate.value, dateFormat)
            QCworksheet.write(QCfileRow, 4, oldinstEfficiencyCell)
            QCworksheet.write(QCfileRow, 5, serialNumber[1])

            QCfileRow += 1

        theFile.close()
        theFile.save(file)

    QCworkbook.close()

    print("The files with no s/n are {}, the sheet is {}".format(filesWithNoMatchingSN, sheetsOfFilesWithNoMatchingSN))
    os.startfile(savePath + '\\' + 'QC.xlsx')

    return
