import openpyxl
import json
import os
import sys
import xlsxwriter


def main(path, savePath):
    if os.path.isfile(savePath):
        os.mkdir(savePath)

    # create excel QC file
    QCworkbook = xlsxwriter.Workbook(savePath + '\\' + 'QC-Removable.xlsx')
    QCworksheet = QCworkbook.add_worksheet()

    # create columns with headers
    QCworksheet.write(0, 0, 'File Name')
    QCworksheet.write(0, 1, 'Sheet Name')
    QCworksheet.write(0, 2, 'Survey Number')
    QCworksheet.write(0, 3, 'Instrument Model')
    QCworksheet.write(0, 4, 'Instrument S/N')
    QCworksheet.write(0, 5, 'Cal DueDate')


    def resource_path(relative_path):
        base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.dirname(__file__)))
        return os.path.join(base_path, relative_path)


    def getListOfFiles(dirName):
        listOfFile = os.listdir(dirName)
        allFiles = list()

        for file in listOfFile:
            fullPath = os.path.join(dirName, file)

            if os.path.isdir(fullPath):
                allFiles = allFiles + getListOfFiles(fullPath)
            else:
                allFiles.append(fullPath)

        print(allFiles)

        return allFiles


    """This function returns the location of the indicated Beta-Gamma"""
    def check_for_BettaGamma(num):
        found = 0
        for row in range(1, 30):
            for column in "GHIJKLMNOPQRSTUVWXYZ":

                modelVal = currentSheet[column +str(row)].value
                if modelVal != "Beta-Gamma":
                    continue

                found += 1
                if found == num:
                    return [row, column]

        return [None, None]

    def second_find_instrument_model_cell(currentSheet):
        for row in range(1, 30):
            for column in "GHIJKLMNOPQRSTUVWXYZ":  # Here you can add or reduce the columns
                modelCell = "{}{}".format(column, row)
                modelVal = currentSheet[modelCell].value

                if (modelVal is None) or isinstance(modelVal, float) == True:
                    continue

                modelVal = str(modelVal)

                if (modelVal[:6] == "ASC-DP") or (modelVal[:4] == "2929") or (modelVal[:4] == "3030"):
                    tempVal = currentSheet[column + str(row+11)].value
                    if tempVal is not None:
                        return [row, column, modelCell, modelVal]

        return [0, 0, None, None]

    def find_instrument_model_cell(currentSheet):
        for row in range(1, 30):
            for column in "GHIJKLMNOPQRSTUVWXYZ":  # Here you can add or reduce the columns
                modelCell = "{}{}".format(column, row)
                modelVal = currentSheet[modelCell].value

                if (modelVal is None) or isinstance(modelVal, float) == True:
                    continue

                modelVal = str(modelVal)

                if (modelVal[:6] == "ASC-DP") or (modelVal[:4] == "2929") or (modelVal[:4] == "3030"):
                    bettarow, gammaRow = check_for_BettaGamma(4)
                    if (bettarow == None):
                        print("bettarow is None")

                    elif currentSheet[gammaRow + str(bettarow+2)].value != None:
                        return [row, column, modelCell, modelVal]
                    else:
                        return [row, column, modelCell, None]

        return [0, 0, None, None]


    def find_instrument_sn_cell(instModelRow, instModelColumn):
        snRow = str(int(instModelRow) + 1)
        snCol = instModelColumn
        snCell = currentSheet[snCol + snRow]

        return snCell


    def find_cal_due_date(instModelRow, instModelColumn):
        calRow = str(int(instModelRow) + 2)
        calCol = chr(ord(instModelColumn))
        calCell = currentSheet[calCol + calRow]

        return calCell


    def find_instrument_efficiency(instModelRow, instModelColumn):
        effRow = str(int(instModelRow) + 3)
        effCol = chr(ord(instModelColumn) + 2)
        effCell = currentSheet[effCol + effRow]

        if type(effCell).__name__ == 'MergedCell':
            effCol = chr(ord(instModelColumn) + 3)
            effCell = currentSheet[effCol + effRow]

        return effCell

    """
    Take the current sheet and returns the value of the cell to the right of the cell containing Survey No.
    """
    def find_survey_number(currentSheet):
        for column in "ABCDEFGHI":
            for row in range(1, 20):
                modelCell = "{}{}".format(column, row)
                newCol = column

                while (currentSheet[modelCell].value == "Survey No") or (currentSheet[modelCell].value == "Survey Number"):
                    if newCol >= "V":
                        continue

                    newCol = chr(ord(newCol) + 1)
                    cell = currentSheet[newCol + str(row)]

                    if type(cell).__name__ != 'MergedCell':
                        newVal = currentSheet[newCol + str(row)].value
                        test = 0
                        return newVal

        return "None"


    def find_efficiency(instSNcell, instEfficiencyCell):
        for inst in instrumentsData:
            if inst['sn'] == instSNcell.value:
                return [inst['sn'], inst['betaEfficiency']]

        return [None, None]


    filesWithNoMatchingSN = list()
    sheetsOfFilesWithNoMatchingSN = list()
    invalidSheets = list()

    files = getListOfFiles(path)

    """This is used for the exe"""
    # with open(file=resource_path("package.json")) as instruments_file:
    #     instrumentsData = json.load(instruments_file)

    """This is used when running the program"""
    with open('package.json') as instruments_file:
        instrumentsData = json.load(instruments_file)

    QCfileRow = 1
    dateFormat = QCworkbook.add_format({'num_format': 'mm/dd/yyyy'})

    for file in files:
        if file[-4:] != "xlsx":
            continue
        try:
            theFile = openpyxl.load_workbook(file)
        except IOError:
            print(file + " could not be opened")
        allSheetNames = theFile.sheetnames

        print("All sheet names {} ".format(theFile.sheetnames))

        for x in allSheetNames:
            print("Current sheet name is {}" .format(x))
            currentSheet = theFile[x]
            instModelRow, instModelColumn, instModelCell, instModel = find_instrument_model_cell(currentSheet)
            surveyNumber = find_survey_number(currentSheet)

            print("The cell is {}, the row is {} and the column is {} ".format(instModelCell, instModelRow, instModelColumn))

            betaRow, betaCol = check_for_BettaGamma(3)
            if betaRow is None or betaCol is None:
                invalidSheets.append(file)
                continue

            if instModelCell is None:
                continue
            instSNcell = find_instrument_sn_cell(instModelRow, instModelColumn)
            instEfficiencyCell = find_instrument_efficiency(instModelRow, instModelColumn)

            instCalDueDate = find_cal_due_date(instModelRow, instModelColumn)
            serialNumber = find_efficiency(instSNcell, instEfficiencyCell)

            if serialNumber[0] is None:
                filesWithNoMatchingSN.append(file)
                sheetsOfFilesWithNoMatchingSN.append(currentSheet)

            # Find the file name
            head, tail = os.path.split(file)

            # Find the Name of the worksheet
            currentSheetString = str(currentSheet)
            currentSheetString = currentSheetString[12:]
            currentSheetString = currentSheetString[:-2]

            # Write the current Worksheet
            QCworksheet.write(QCfileRow, 0, tail)
            QCworksheet.write(QCfileRow, 1, currentSheetString)
            QCworksheet.write(QCfileRow, 2, surveyNumber)
            QCworksheet.write(QCfileRow, 3, instModel)
            QCworksheet.write(QCfileRow, 4, instSNcell.value)
            QCworksheet.write(QCfileRow, 5, instCalDueDate.value, dateFormat)

            QCfileRow += 1


        theFile.close()


    """redo but for the other list"""
    print("IN THE INVALID BETA ")
    print(invalidSheets)

    for file in invalidSheets:
        theFile = openpyxl.load_workbook(file)
        allSheetNames = theFile.sheetnames

        for x in allSheetNames:
            print("Current sheet name is {}" .format(x))
            currentSheet = theFile[x]
            instModelRow, instModelColumn, instModelCell, instModel = second_find_instrument_model_cell(currentSheet)
            surveyNumber = find_survey_number(currentSheet)

            print(instModelCell)

            if instModelCell is None:
                continue

            instSNcell = find_instrument_sn_cell(instModelRow, instModelColumn)
            instEfficiencyCell = find_instrument_efficiency(instModelRow, instModelColumn)

            instCalDueDate = find_cal_due_date(instModelRow, instModelColumn)
            serialNumber = find_efficiency(instSNcell, instEfficiencyCell)

            if serialNumber[0] is None:
                filesWithNoMatchingSN.append(file)
                sheetsOfFilesWithNoMatchingSN.append(currentSheet)

            # Find the file name
            head, tail = os.path.split(file)

            # Find the Name of the worksheet
            currentSheetString = str(currentSheet)
            currentSheetString = currentSheetString[12:]
            currentSheetString = currentSheetString[:-2]

            # Write the current Worksheet
            QCworksheet.write(QCfileRow, 0, tail)
            QCworksheet.write(QCfileRow, 1, currentSheetString)
            QCworksheet.write(QCfileRow, 2, surveyNumber)
            QCworksheet.write(QCfileRow, 3, instModel)
            QCworksheet.write(QCfileRow, 4, instSNcell.value)
            QCworksheet.write(QCfileRow, 5, instCalDueDate.value, dateFormat)

            QCfileRow += 1

        theFile.close()

    QCworkbook.close()

    print("The files with no s/n are {}, the sheet is {}".format(filesWithNoMatchingSN, sheetsOfFilesWithNoMatchingSN))
    os.startfile(savePath + '\\' + 'QC-Removable.xlsx')

    return


